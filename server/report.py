"""Report generation module — query, Excel, email."""

import json
import logging
import os
import smtplib
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from io import BytesIO

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

import config
import models

logger = logging.getLogger(__name__)

# ── Data Query ──────────────────────────────────────────────────────────────


def query_report_data(since: datetime) -> tuple[list[dict], list[dict]]:
    """Query products and reviews added since the given timestamp (Asia/Shanghai).

    Returns (products, reviews) — both as lists of dicts.
    """
    since_str = since.strftime("%Y-%m-%d %H:%M:%S")

    conn = models.get_conn()
    try:
        product_rows = conn.execute(
            """
            SELECT url, name, sku, price, stock_status, rating, review_count,
                   scraped_at, site, ownership
            FROM products
            WHERE scraped_at >= ?
            ORDER BY scraped_at DESC
            """,
            (since_str,),
        ).fetchall()
        products = [dict(r) for r in product_rows]

        review_rows = conn.execute(
            """
            SELECT p.name AS product_name,
                   r.author, r.headline, r.body, r.rating,
                   r.date_published, r.images, p.ownership,
                   r.headline_cn, r.body_cn, r.translate_status
            FROM reviews r
            JOIN products p ON r.product_id = p.id
            WHERE r.scraped_at >= ?
            ORDER BY r.scraped_at DESC
            """,
            (since_str,),
        ).fetchall()
        reviews = []
        for row in review_rows:
            d = dict(row)
            # Normalise images: may be JSON string or None
            if d.get("images") and isinstance(d["images"], str):
                try:
                    d["images"] = json.loads(d["images"])
                except Exception:
                    pass
            reviews.append(d)
    finally:
        conn.close()

    logger.info(
        "query_report_data: since=%s → %d products, %d reviews",
        since_str,
        len(products),
        len(reviews),
    )
    return products, reviews



# ── Excel Generation ─────────────────────────────────────────────────────────


def _cell_value(v):
    """Convert a value to something Excel can store."""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


_IMG_THUMB_HEIGHT = 80   # 缩略图高度（像素）
_IMG_THUMB_SPACING = 5   # 多张图片间距（像素）
_IMG_COL_WIDTH = 17      # 照片列宽度（字符数，约 120px）
_IMG_DOWNLOAD_TIMEOUT = 10  # 单张图片下载超时（秒）


def _download_and_resize(url: str) -> XlImage | None:
    """Download an image and return an openpyxl Image with thumbnail display size.

    Keeps original resolution for clarity when zoomed in;
    only sets the display dimensions to thumbnail size in Excel.
    """
    try:
        resp = requests.get(url, timeout=_IMG_DOWNLOAD_TIMEOUT)
        resp.raise_for_status()
        buf = BytesIO(resp.content)
        # Read original dimensions for aspect ratio calculation
        img = PILImage.open(buf)
        ratio = _IMG_THUMB_HEIGHT / img.height
        display_width = int(img.width * ratio)
        # Create openpyxl Image from original bytes (no pixel resize)
        buf.seek(0)
        xl_img = XlImage(buf)
        # Set display size only — original resolution preserved
        xl_img.width = display_width
        xl_img.height = _IMG_THUMB_HEIGHT
        return xl_img
    except Exception as exc:
        logger.warning("_download_and_resize: failed for %s — %s", url[:80], exc)
        return None


def generate_excel(
    products: list[dict],
    reviews: list[dict],
    report_date: datetime | None = None,
) -> str:
    """Generate an Excel report and return the file path.

    Creates ``config.REPORT_DIR/scrape-report-YYYY-MM-DD.xlsx``.
    Empty data still produces a valid file with headers.
    Review images are downloaded and embedded as thumbnails in cells.
    """
    if report_date is None:
        report_date = config.now_shanghai()

    os.makedirs(config.REPORT_DIR, exist_ok=True)
    filename = f"scrape-report-{report_date.strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(config.REPORT_DIR, filename)

    wb = Workbook()

    # ── Header styling helpers ────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_sheet(ws, headers: list[str], rows: list[list]):
        # Write header
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data rows
        for row_data in rows:
            ws.append([_cell_value(v) for v in row_data])

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    # ── 产品 sheet ────────────────────────────────────
    ws_products = wb.active
    ws_products.title = "产品"
    product_headers = [
        "产品地址", "产品名称", "SKU", "售价$", "库存状态",
        "综合评分", "评论数量", "抓取时间", "站点", "归属",
    ]
    product_keys = [
        "url", "name", "sku", "price", "stock_status",
        "rating", "review_count", "scraped_at", "site", "ownership",
    ]
    product_rows = [[p.get(k) for k in product_keys] for p in products]
    _write_sheet(ws_products, product_headers, product_rows)

    # ── 评论 sheet ────────────────────────────────────
    ws_reviews = wb.create_sheet("评论")
    review_headers = [
        "产品名称", "评论人", "标题（原文）", "内容（原文）",
        "标题（中文）", "内容（中文）", "打分", "评论时间", "照片",
    ]
    # Write header
    ws_reviews.append(review_headers)
    for col_idx, _ in enumerate(review_headers, start=1):
        cell = ws_reviews.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    images_col = len(review_headers)  # "照片" is the last column
    images_col_letter = get_column_letter(images_col)

    # Write review rows with embedded images
    for row_idx, r in enumerate(reviews, start=2):
        ws_reviews.cell(row=row_idx, column=1, value=_cell_value(r.get("product_name")))
        ws_reviews.cell(row=row_idx, column=2, value=_cell_value(r.get("author")))
        ws_reviews.cell(row=row_idx, column=3, value=_cell_value(r.get("headline")))
        ws_reviews.cell(row=row_idx, column=4, value=_cell_value(r.get("body")))
        ws_reviews.cell(row=row_idx, column=5, value=_cell_value(r.get("headline_cn")))
        ws_reviews.cell(row=row_idx, column=6, value=_cell_value(r.get("body_cn")))
        ws_reviews.cell(row=row_idx, column=7, value=_cell_value(r.get("rating")))
        ws_reviews.cell(row=row_idx, column=8, value=_cell_value(r.get("date_published")))

        # Embed images as thumbnails
        image_urls = r.get("images") or []
        if isinstance(image_urls, str):
            try:
                image_urls = json.loads(image_urls)
            except Exception:
                image_urls = []

        embedded_count = 0
        for img_idx, url in enumerate(image_urls):
            xl_img = _download_and_resize(url)
            if xl_img:
                y_offset = img_idx * (_IMG_THUMB_HEIGHT + _IMG_THUMB_SPACING)
                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                from openpyxl.utils.units import pixels_to_EMU
                marker = AnchorMarker(
                    col=images_col - 1,  # 0-indexed
                    row=row_idx - 1,     # 0-indexed
                    colOff=0,
                    rowOff=pixels_to_EMU(y_offset),
                )
                size = XDRPositiveSize2D(
                    pixels_to_EMU(xl_img.width),
                    pixels_to_EMU(xl_img.height),
                )
                xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws_reviews.add_image(xl_img)
                embedded_count += 1

        if embedded_count > 0:
            # Adjust row height to fit stacked images
            row_height_px = embedded_count * (_IMG_THUMB_HEIGHT + _IMG_THUMB_SPACING)
            ws_reviews.row_dimensions[row_idx].height = row_height_px * 0.75  # px to points
        elif image_urls:
            # Fallback: write URLs as text (only if non-empty)
            ws_reviews.cell(row=row_idx, column=images_col, value=_cell_value(image_urls))

    # Auto-adjust non-image column widths
    for col in ws_reviews.columns:
        col_letter = col[0].column_letter
        if col_letter == images_col_letter:
            ws_reviews.column_dimensions[col_letter].width = _IMG_COL_WIDTH
            continue
        max_len = 0
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws_reviews.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    wb.save(filepath)
    logger.info("generate_excel: saved to %s", filepath)
    return filepath


# ── Email ────────────────────────────────────────────────────────────────────


def load_email_recipients(filepath: str) -> list[str]:
    """Read email addresses from a file, one per line.

    Lines starting with ``#`` (after stripping) are treated as comments.
    Returns a list of non-empty email addresses.
    """
    recipients: list[str] = []
    try:
        with open(filepath, encoding="utf-8") as f:
            for line in f:
                stripped = line.strip()
                if stripped and not stripped.startswith("#"):
                    recipients.append(stripped)
    except FileNotFoundError:
        logger.warning("load_email_recipients: file not found — %s", filepath)
    except Exception as exc:
        logger.warning("load_email_recipients: error reading %s — %s", filepath, exc)
    return recipients


def send_email(
    recipients: list[str],
    subject: str,
    body_text: str,
    attachment_path: str | None = None,
) -> dict:
    """Send an email via SMTP.

    Returns a dict with keys:
    - success (bool)
    - error (str | None)
    - recipients (int) — number of addresses the message was sent to
    """
    if not recipients:
        return {"success": False, "error": "No recipients provided", "recipients": 0}

    if not config.SMTP_HOST:
        return {"success": False, "error": "SMTP_HOST not configured", "recipients": 0}

    try:
        msg = MIMEMultipart()
        msg["From"] = config.SMTP_FROM or config.SMTP_USER
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        if attachment_path and os.path.isfile(attachment_path):
            with open(attachment_path, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(attachment_path)}",
            )
            msg.attach(part)

        if config.SMTP_USE_SSL:
            smtp = smtplib.SMTP_SSL(config.SMTP_HOST, config.SMTP_PORT)
        else:
            smtp = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT)
            smtp.starttls()

        if config.SMTP_USER and config.SMTP_PASSWORD:
            smtp.login(config.SMTP_USER, config.SMTP_PASSWORD)

        smtp.sendmail(msg["From"], recipients, msg.as_string())
        smtp.quit()

        logger.info("send_email: sent to %d recipients — %s", len(recipients), subject)
        return {"success": True, "error": None, "recipients": len(recipients)}

    except Exception as exc:
        logger.warning("send_email: failed — %s", exc)
        return {"success": False, "error": str(exc), "recipients": 0}


# ── Pipeline ─────────────────────────────────────────────────────────────────

# Alias before defining generate_report so the parameter name doesn't shadow it
_send_email_impl = send_email


def generate_report(
    since: datetime,
    send_email: bool = True,
) -> dict:
    """Report pipeline: query (with pre-translated data) → Excel → email.

    Translation is handled by the background TranslationWorker.
    Reviews that haven't been translated yet will have empty Chinese fields.
    """
    # 1. Query (includes headline_cn/body_cn from DB)
    products, reviews = query_report_data(since)

    # 2. Count translation status
    translated_count = sum(
        1 for r in reviews if r.get("translate_status") == "done"
    )
    untranslated_count = len(reviews) - translated_count

    # Ensure Chinese fields exist for Excel generation
    for r in reviews:
        r.setdefault("headline_cn", "")
        r.setdefault("body_cn", "")

    # 3. Excel
    excel_path = generate_excel(products, reviews, report_date=since)

    # 4. Email
    email_result = None
    if send_email:
        recipients = config.EMAIL_RECIPIENTS
        since_str = since.strftime("%Y-%m-%d")
        subject = f"Qbu 每日抓取报告 {since_str}"
        body = (
            f"您好，\n\n"
            f"以下是 {since_str} 的 Qbu 产品抓取报告汇总：\n"
            f"  - 产品数：{len(products)}\n"
            f"  - 评论数：{len(reviews)}\n"
            f"  - 已翻译评论：{translated_count}\n"
        )
        if untranslated_count > 0:
            body += f"  - 注：{untranslated_count} 条评论翻译进行中，中文列暂时为空\n"
        body += f"\n详细数据请查阅附件 Excel 文件。\n"
        email_result = _send_email_impl(
            recipients=recipients,
            subject=subject,
            body_text=body,
            attachment_path=excel_path,
        )

    return {
        "products_count": len(products),
        "reviews_count": len(reviews),
        "translated_count": translated_count,
        "untranslated_count": untranslated_count,
        "excel_path": excel_path,
        "email": email_result,
    }
