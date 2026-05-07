"""Report generation module — query, Excel, email."""

import json
import logging
import os
import smtplib
import time
from dataclasses import asdict
from datetime import datetime, timedelta, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from io import BytesIO

import requests
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ── Conditional formatting fills for analytical sheets ────────────────────────
_RED_FILL = PatternFill(start_color="F2D9D0", end_color="F2D9D0", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="F5ECD4", end_color="F5ECD4", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="DCE9E3", end_color="DCE9E3", fill_type="solid")

from qbu_crawler import config, models
from qbu_crawler.server.report_common import (  # noqa: F401 — re-export
    _LABEL_DISPLAY,
    _PRIORITY_DISPLAY,
    _SEVERITY_DISPLAY,
    _derive_review_label_codes,
    _fallback_executive_bullets,
    _fallback_hero_headline,
    _join_label_codes,
    _join_label_counts,
    _label_display,
    _summary_text,
    normalize_deep_report_analytics,
)
from qbu_crawler.server.scope import Scope, normalize_scope

logger = logging.getLogger(__name__)

# ── F011 §6.4.1 v1.1 — REPORT_TEMPLATE_VERSION env routing ───────────────────
REPORT_TEMPLATE_DIR = Path(__file__).parent / "report_templates"
_template_log = logging.getLogger(__name__)


def _select_template(env_template_version: str | None = None) -> str:
    """F011 §6.4.1 — Select attachment HTML template based on env var.

    REPORT_TEMPLATE_VERSION:
      - "v3" (default) → daily_report_v3.html.j2
      - "v3_legacy" → daily_report_v3_legacy.html.j2 (rollback path)
      - any other value → fallback to v3 + WARNING
      - if mapped file missing → fallback to v3 + WARNING
    """
    version = env_template_version or os.environ.get("REPORT_TEMPLATE_VERSION", "v3")
    template_map = {
        "v3": "daily_report_v3.html.j2",
        "v3_legacy": "daily_report_v3_legacy.html.j2",
    }
    if version not in template_map:
        _template_log.warning(
            "Unknown REPORT_TEMPLATE_VERSION=%s, fallback to v3", version
        )
        return template_map["v3"]
    template_file = template_map[version]
    template_path = REPORT_TEMPLATE_DIR / template_file
    if not template_path.exists():
        _template_log.warning(
            "Template %s missing, fallback to v3", template_path
        )
        return template_map["v3"]
    return template_file


_SMTP_RETRY_ATTEMPTS = 3

# ── Data Query ──────────────────────────────────────────────────────────────


def query_report_data(since: datetime) -> tuple[list[dict], list[dict]]:
    """Query products and reviews added since the given timestamp (Asia/Shanghai).

    Returns (products, reviews) — both as lists of dicts.
    """
    since_str = since.strftime("%Y-%m-%d %H:%M:%S")

    conn = models.get_conn()
    try:
        product_columns = _product_select_columns(conn)
        product_rows = conn.execute(
            f"""
            SELECT {product_columns}
            FROM products
            WHERE scraped_at >= ?
            ORDER BY scraped_at DESC
            """,
            (since_str,),
        ).fetchall()
        products = [dict(r) for r in product_rows]

        review_rows = conn.execute(
            """
            SELECT r.id AS id, p.name AS product_name, p.sku AS product_sku,
                   p.url AS product_url, p.site AS site, r.scraped_at AS scraped_at,
                   r.author, r.headline, r.body, r.rating,
                   r.date_published, r.date_published_parsed, r.images,
                   p.ownership,
                   r.headline_cn, r.body_cn, r.translate_status
            FROM reviews r
            JOIN products p ON r.product_id = p.id
            WHERE r.scraped_at >= ?
            ORDER BY r.scraped_at DESC
            """,
            (since_str,),
        ).fetchall()
        reviews = []
        for row in review_rows:
            d = dict(row)
            # Normalise images: may be JSON string or None
            if d.get("images") and isinstance(d["images"], str):
                try:
                    d["images"] = json.loads(d["images"])
                except Exception:
                    pass
            reviews.append(d)
    finally:
        conn.close()

    logger.info(
        "query_report_data: since=%s → %d products, %d reviews",
        since_str,
        len(products),
        len(reviews),
    )
    return products, reviews



# ── Excel Generation ─────────────────────────────────────────────────────────


def _cell_value(v):
    """Convert a value to something Excel can store."""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


# ── Excel display-name mappings (DB values → Chinese for human readers) ──────
# These are terminal display-only; original English values remain in DB/analytics.
_OWNERSHIP_DISPLAY = {"own": "自有", "competitor": "竞品"}
_SENTIMENT_DISPLAY = {"positive": "正面", "negative": "负面", "mixed": "复杂", "neutral": "中性"}
_POLARITY_DISPLAY = {"positive": "正面", "negative": "负面"}
_SEVERITY_DISPLAY = {"critical": "危急", "high": "高", "medium": "中", "low": "低"}
_STOCK_DISPLAY = {"in_stock": "有货", "out_of_stock": "缺货", "unknown": "未知"}

# F011 §4.3 — impact_category enum (severity-axis) display
IMPACT_CATEGORY_DISPLAY = {
    "functional": "功能性",
    "durability": "耐用性",
    "safety": "安全性",
    "cosmetic": "外观",
    "service": "服务",
}

# F011 §4.3 + H19 — failure_mode 9-class enum display
FAILURE_MODE_DISPLAY = {
    "none": "无",
    "gear_failure": "齿轮失效",
    "motor_anomaly": "电机异常",
    "casing_assembly": "壳体/装配",
    "material_finish": "表面/材料",
    "control_electrical": "控制/电气",
    "noise": "噪音",
    "cleaning_difficulty": "清洁困难",
    "other": "其他",
}

# F011 §4.3 — status lamp icon prefix for the 核心数据 sheet
_STATUS_LAMP_DISPLAY = {
    "red": "🔴 高风险",
    "yellow": "🟡 需关注",
    "green": "🟢 良好",
    "gray": "⚪ 无数据",
}

from qbu_crawler.server.report_common import _LABEL_DISPLAY  # label_code → 中文


def _xl_display(value, mapping: dict) -> str:
    """Map a raw value to its Chinese display string for Excel output."""
    if value is None:
        return ""
    return mapping.get(value, value)


_IMG_THUMB_HEIGHT = 80   # 缩略图高度（像素）
_IMG_THUMB_SPACING = 5   # 多张图片间距（像素）
_IMG_COL_WIDTH = 17      # 照片列宽度（字符数，约 120px）
_IMG_DOWNLOAD_TIMEOUT = 10  # 单张图片下载超时（秒）
_IMG_MAX_BYTES = 5 * 1024 * 1024  # 单张图片最大 5 MB，超过跳过

# Sentinel to distinguish "download failed" from "not in cache" so that
# failed URLs are not retried on every row that references them.
_IMG_FAILED = "FAILED"


def _download_image_data(url: str) -> tuple[bytes, int, int] | None:
    """Download an image and return (raw_bytes, display_width, display_height).

    Returns None on failure. The raw bytes are kept so each call site can create
    its own openpyxl Image — a single XlImage instance must not be added to
    multiple anchors.
    """
    try:
        for _attempt in range(2):
            try:
                resp = requests.get(url, timeout=_IMG_DOWNLOAD_TIMEOUT)
                resp.raise_for_status()
                break
            except requests.RequestException:
                if _attempt == 1:
                    raise
        raw = resp.content
        if not raw:
            return None
        if len(raw) > _IMG_MAX_BYTES:
            logger.warning("_download_image_data: skipping oversized image (%d bytes): %s",
                           len(raw), url[:80])
            return None
        # Read original dimensions for aspect ratio calculation
        img = PILImage.open(BytesIO(raw))
        orig_w, orig_h = img.size
        img.close()
        if orig_h <= 0 or orig_w <= 0:
            return None
        ratio = _IMG_THUMB_HEIGHT / orig_h
        display_width = int(orig_w * ratio)
        return raw, display_width, _IMG_THUMB_HEIGHT
    except Exception as exc:
        logger.warning("_download_image_data: failed for %s — %s", url[:80], exc)
        return None


def _make_xl_image(data: tuple[bytes, int, int]) -> XlImage:
    """Create a fresh XlImage from cached (raw_bytes, width, height)."""
    raw, w, h = data
    xl_img = XlImage(BytesIO(raw))
    xl_img.width = w
    xl_img.height = h
    return xl_img


def _download_images_parallel(urls: list[str], global_timeout: float = 60) -> dict:
    """Download multiple images in parallel.

    Returns {url: (bytes,w,h) | _IMG_FAILED}.  Every input URL is guaranteed
    to have an entry so callers can distinguish "download failed" from "URL
    not in cache" via the _IMG_FAILED sentinel.
    """
    if not urls:
        return {}

    from concurrent.futures import ThreadPoolExecutor, as_completed

    results = {}
    with ThreadPoolExecutor(max_workers=5) as pool:
        future_to_url = {pool.submit(_download_image_data, url): url for url in urls}
        try:
            for future in as_completed(future_to_url, timeout=global_timeout):
                url = future_to_url[future]
                try:
                    results[url] = future.result() or _IMG_FAILED
                except Exception:
                    results[url] = _IMG_FAILED
        except TimeoutError:
            # Global timeout reached — mark remaining as failed
            for future, url in future_to_url.items():
                if url not in results:
                    results[url] = _IMG_FAILED
                    future.cancel()

    return results


def generate_excel(
    products: list[dict],
    reviews: list[dict],
    report_date: datetime | None = None,
) -> str:
    """Generate an Excel report and return the file path.

    Creates ``config.REPORT_DIR/scrape-report-YYYY-MM-DD.xlsx``.
    Empty data still produces a valid file with headers.
    Review images are downloaded and embedded as thumbnails in cells.
    """
    if report_date is None:
        report_date = config.now_shanghai()
    if report_date.tzinfo is not None:
        report_date = report_date.astimezone(config.SHANGHAI_TZ).replace(tzinfo=None)

    os.makedirs(config.REPORT_DIR, exist_ok=True)
    filename = f"scrape-report-{report_date.strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(config.REPORT_DIR, filename)

    wb = Workbook()

    # ── Header styling helpers ────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_sheet(ws, headers: list[str], rows: list[list]):
        # Write header
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data rows
        for row_data in rows:
            ws.append([_cell_value(v) for v in row_data])

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    # ── 产品 sheet ────────────────────────────────────
    ws_products = wb.active
    ws_products.title = "产品"
    product_headers = [
        "产品地址", "产品名称", "SKU", "售价$", "库存状态",
        "综合评分", "评论数量", "抓取时间", "站点", "归属",
    ]
    product_keys = [
        "url", "name", "sku", "price", "stock_status",
        "rating", "review_count", "scraped_at", "site", "ownership",
    ]
    product_rows = [[p.get(k) for k in product_keys] for p in products]
    _write_sheet(ws_products, product_headers, product_rows)

    # ── 评论 sheet ────────────────────────────────────
    ws_reviews = wb.create_sheet("评论")
    review_headers = [
        "产品名称", "SKU", "评论人", "标题（原文）", "内容（原文）",
        "标题（中文）", "内容（中文）", "打分", "评论时间", "照片",
    ]
    # Write header
    ws_reviews.append(review_headers)
    for col_idx, _ in enumerate(review_headers, start=1):
        cell = ws_reviews.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    images_col = len(review_headers)  # "照片" is the last column
    images_col_letter = get_column_letter(images_col)

    # Pre-fetch all review images in parallel
    _all_image_urls = set()
    for _r in reviews:
        _imgs = _r.get("images") or []
        if isinstance(_imgs, str):
            try:
                _imgs = json.loads(_imgs)
            except Exception:
                _imgs = []
        if isinstance(_imgs, list):
            for _url in _imgs:
                if isinstance(_url, str) and _url.startswith("http"):
                    _all_image_urls.add(_url)
    _prefetched = _download_images_parallel(list(_all_image_urls)) if _all_image_urls else {}

    # Write review rows with embedded images
    for row_idx, r in enumerate(reviews, start=2):
        ws_reviews.cell(row=row_idx, column=1, value=_cell_value(r.get("product_name")))
        ws_reviews.cell(row=row_idx, column=2, value=_cell_value(r.get("product_sku")))
        ws_reviews.cell(row=row_idx, column=3, value=_cell_value(r.get("author")))
        ws_reviews.cell(row=row_idx, column=4, value=_cell_value(r.get("headline")))
        ws_reviews.cell(row=row_idx, column=5, value=_cell_value(r.get("body")))
        ws_reviews.cell(row=row_idx, column=6, value=_cell_value(r.get("headline_cn")))
        ws_reviews.cell(row=row_idx, column=7, value=_cell_value(r.get("body_cn")))
        ws_reviews.cell(row=row_idx, column=8, value=_cell_value(r.get("rating")))
        ws_reviews.cell(row=row_idx, column=9, value=_cell_value(r.get("date_published")))

        # Embed images as thumbnails
        image_urls = r.get("images") or []
        if isinstance(image_urls, str):
            try:
                image_urls = json.loads(image_urls)
            except Exception:
                image_urls = []

        embedded_count = 0
        for url in image_urls:
            if not isinstance(url, str) or not url.startswith("http"):
                continue
            cached = _prefetched.get(url, _IMG_FAILED)
            if cached is _IMG_FAILED:
                cached = _download_image_data(url)
            if not cached or cached is _IMG_FAILED:
                continue
            xl_img = _make_xl_image(cached)
            y_offset = embedded_count * (_IMG_THUMB_HEIGHT + _IMG_THUMB_SPACING)
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            marker = AnchorMarker(
                col=images_col - 1,  # 0-indexed
                row=row_idx - 1,     # 0-indexed
                colOff=0,
                rowOff=pixels_to_EMU(y_offset),
            )
            size = XDRPositiveSize2D(
                pixels_to_EMU(xl_img.width),
                pixels_to_EMU(xl_img.height),
            )
            xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws_reviews.add_image(xl_img)
            embedded_count += 1

        if embedded_count > 0:
            # Adjust row height to fit stacked images
            row_height_px = embedded_count * (_IMG_THUMB_HEIGHT + _IMG_THUMB_SPACING)
            ws_reviews.row_dimensions[row_idx].height = row_height_px * 0.75  # px to points
        elif image_urls:
            # Fallback: write URLs as text (only if non-empty)
            ws_reviews.cell(row=row_idx, column=images_col, value=_cell_value(image_urls))

    # Auto-adjust non-image column widths
    for col in ws_reviews.columns:
        col_letter = col[0].column_letter
        if col_letter == images_col_letter:
            ws_reviews.column_dimensions[col_letter].width = _IMG_COL_WIDTH
            continue
        max_len = 0
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws_reviews.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    wb.save(filepath)
    logger.info("generate_excel: saved to %s", filepath)
    return filepath


# ── Email ────────────────────────────────────────────────────────────────────


def load_email_recipients(filepath: str) -> list[str]:
    """Read email addresses from a file, one per line.

    Lines starting with ``#`` (after stripping) are treated as comments.
    Returns a list of non-empty email addresses.
    """
    recipients: list[str] = []
    try:
        with open(filepath, encoding="utf-8") as f:
            for line in f:
                stripped = line.strip()
                if stripped and not stripped.startswith("#"):
                    recipients.append(stripped)
    except FileNotFoundError:
        logger.warning("load_email_recipients: file not found — %s", filepath)
    except Exception as exc:
        logger.warning("load_email_recipients: error reading %s — %s", filepath, exc)
    return recipients


def send_email(
    recipients: list[str],
    subject: str,
    body_text: str,
    body_html: str | None = None,
    attachment_path: str | None = None,
    attachment_paths: list[str] | None = None,
) -> dict:
    """Send an email via SMTP.

    Returns a dict with keys:
    - success (bool)
    - error (str | None)
    - recipients (int) — number of addresses the message was sent to
    """
    if not recipients:
        return {"success": False, "error": "No recipients provided", "recipients": 0}

    if not config.SMTP_HOST:
        return {"success": False, "error": "SMTP_HOST not configured", "recipients": 0}

    attachments = [path for path in (attachment_paths or []) if path]
    if attachment_path:
        attachments.insert(0, attachment_path)

    last_error = None
    for attempt in range(_SMTP_RETRY_ATTEMPTS):
        smtp = None
        try:
            msg = MIMEMultipart("mixed")
            msg["From"] = config.SMTP_FROM or config.SMTP_USER
            if config.EMAIL_BCC_MODE:
                msg["To"] = config.SMTP_FROM or config.SMTP_USER
                msg["Bcc"] = ", ".join(recipients)
            else:
                msg["To"] = ", ".join(recipients)
            msg["Subject"] = subject

            body_part = MIMEMultipart("alternative")
            body_part.attach(MIMEText(body_text, "plain", "utf-8"))
            if body_html:
                body_part.attach(MIMEText(body_html, "html", "utf-8"))
            msg.attach(body_part)

            for path in attachments:
                if not os.path.isfile(path):
                    continue
                with open(path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    "attachment",
                    filename=("utf-8", "", os.path.basename(path)),
                )
                msg.attach(part)

            if config.SMTP_USE_SSL:
                smtp = smtplib.SMTP_SSL(config.SMTP_HOST, config.SMTP_PORT)
            else:
                smtp = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT)
                smtp.starttls()

            if config.SMTP_USER and config.SMTP_PASSWORD:
                smtp.login(config.SMTP_USER, config.SMTP_PASSWORD)

            smtp.sendmail(msg["From"], recipients, msg.as_string())
            logger.info("send_email: sent to %d recipients — %s", len(recipients), subject)
            return {"success": True, "error": None, "recipients": len(recipients)}
        except Exception as exc:
            last_error = exc
            logger.warning(
                "send_email: attempt %d/%d failed — %s",
                attempt + 1,
                _SMTP_RETRY_ATTEMPTS,
                exc,
            )
            if attempt + 1 < _SMTP_RETRY_ATTEMPTS:
                time.sleep(2**attempt)
        finally:
            if smtp is not None:
                try:
                    smtp.quit()
                except Exception:
                    close = getattr(smtp, "close", None)
                    if callable(close):
                        try:
                            close()
                        except Exception:
                            pass

    return {"success": False, "error": str(last_error), "recipients": 0}


def _report_template_dir():
    return Path(__file__).with_name("report_templates")


def _report_template_env():
    return Environment(
        loader=FileSystemLoader(str(_report_template_dir())),
        autoescape=False,
        trim_blocks=True,
        lstrip_blocks=True,
    )


def render_daily_email_html(snapshot, analytics):
    """Render the HTML email template."""
    normalized = normalize_deep_report_analytics(analytics)
    _ensure_humanized_bullets(normalized)
    env = _report_template_env()
    template = env.get_template("daily_report_email.html.j2")
    return template.render(
        snapshot=snapshot,
        analytics=normalized,
        threshold=config.NEGATIVE_THRESHOLD,
    )


def render_email_full(snapshot, analytics):
    """F011 §4.1 — render the full-mode email body (email_full.html.j2).

    Public entry point. The new template only consumes:
      - logical_date  (from snapshot)
      - analytics.kpis.{health_index, own_positive_review_rows, own_review_rows,
                        own_negative_review_rate, own_negative_review_rate_display,
                        own_product_count, competitor_product_count,
                        ingested_review_rows}
      - analytics.report_user_contract.action_priorities[]
      - analytics.report_copy.{hero_headline, executive_bullets[]}
      - analytics.self.product_status[].{product_name, status_lamp,
                                          primary_concern}  (Task 3.4)

    Other legacy template vars (change_digest, risk_products, alert_level,
    cumulative_kpis, window, etc.) are no longer consumed by §4.1; the new
    layout deliberately drops them.
    """
    from jinja2 import select_autoescape

    template_dir = Path(__file__).parent / "report_templates"
    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=select_autoescape(["html", "j2"]),
    )

    from qbu_crawler.server.report_contract import build_report_user_contract

    email_source = dict(analytics or {})
    input_contract = email_source.get("report_user_contract") or {}
    if not email_source.get("kpis") and input_contract.get("kpis"):
        email_source["kpis"] = dict(input_contract.get("kpis") or {})
    input_report_copy = email_source.get("report_copy") or {}
    input_kpis = dict(email_source.get("kpis") or {})

    email_analytics = normalize_deep_report_analytics(email_source)
    if "health_index" in input_kpis:
        email_analytics.setdefault("kpis", {})["health_index"] = input_kpis["health_index"]
    if "hero_headline" in input_report_copy:
        email_analytics.setdefault("report_copy", {})["hero_headline"] = input_report_copy.get("hero_headline") or ""
    if "executive_bullets" in input_report_copy:
        email_analytics.setdefault("report_copy", {})["executive_bullets"] = input_report_copy.get("executive_bullets") or []
    contract = email_analytics.get("report_user_contract") or {}
    if (contract.get("contract_context") or {}).get("snapshot_source") != "provided":
        email_analytics["report_user_contract"] = build_report_user_contract(
            snapshot=snapshot or {},
            analytics=email_analytics,
        )
    if "health_index" in input_kpis:
        email_analytics.setdefault("report_user_contract", {}).setdefault("kpis", {})["health_index"] = input_kpis["health_index"]
    email_analytics["email_evidence"] = _build_email_evidence(snapshot or {})
    tpl = env.get_template("email_full.html.j2")
    return tpl.render(
        snapshot=snapshot or {},
        logical_date=snapshot.get("logical_date", "") if snapshot else "",
        analytics=email_analytics,
    )


def _build_email_evidence(snapshot):
    reviews = list((snapshot or {}).get("reviews") or [])

    def rating_value(review, default=0):
        try:
            return float(review.get("rating"))
        except (TypeError, ValueError):
            return default

    def parse_labels(review):
        labels = review.get("analysis_labels")
        if isinstance(labels, str):
            try:
                labels = json.loads(labels)
            except Exception:
                labels = []
        if not isinstance(labels, list):
            return []
        return [item for item in labels if isinstance(item, dict)]

    def label_display(review):
        labels = parse_labels(review)
        if labels:
            first = labels[0]
            code = str(first.get("code") or "").strip()
            return (
                str(first.get("display") or first.get("label") or "").strip()
                or _LABEL_DISPLAY.get(code)
                or code
            )
        return str(review.get("impact_category") or review.get("failure_mode") or "").strip()

    def truncate(value, limit=180):
        text = str(value or "").replace("\n", " ").strip()
        if len(text) <= limit:
            return text
        return text[: limit - 1].rstrip() + "…"

    def item(review):
        return {
            "sku": review.get("product_sku") or review.get("sku") or "未知 SKU",
            "rating": review.get("rating"),
            "issue": label_display(review) or "未分类",
            "original": truncate(review.get("body") or review.get("headline")),
            "translation": truncate(review.get("body_cn") or review.get("headline_cn")) or "翻译中",
            "analysis": truncate(review.get("analysis_insight_cn") or review.get("analysis_insight_en") or label_display(review), 120),
        }

    own = [r for r in reviews if r.get("ownership") == "own"]
    competitors = [r for r in reviews if r.get("ownership") == "competitor"]
    own_risk = sorted(
        [r for r in own if rating_value(r, 5) <= config.NEGATIVE_THRESHOLD],
        key=lambda r: (rating_value(r, 5), str(r.get("scraped_at") or "")),
    )
    own_positive = sorted(
        [r for r in own if rating_value(r, 0) >= 4 or str(r.get("sentiment") or "").lower() == "positive"],
        key=lambda r: (-rating_value(r, 0), str(r.get("scraped_at") or "")),
    )
    competitor_positive = sorted(
        [r for r in competitors if rating_value(r, 0) >= 4 or str(r.get("sentiment") or "").lower() == "positive"],
        key=lambda r: (-rating_value(r, 0), str(r.get("scraped_at") or "")),
    )
    competitor_negative = sorted(
        [r for r in competitors if rating_value(r, 5) <= config.NEGATIVE_THRESHOLD],
        key=lambda r: (rating_value(r, 5), str(r.get("scraped_at") or "")),
    )
    own_source = own_risk or own_positive or own
    competitor_source = competitor_positive or competitor_negative or competitors
    own_title = "自有风险证据" if own_risk else ("自有亮点证据" if own_positive else "自有新增评论证据")
    competitor_title = (
        "竞品亮点证据" if competitor_positive
        else ("竞品机会证据" if competitor_negative else "竞品新增评论证据")
    )
    return {
        "own_title": own_title,
        "competitor_title": competitor_title,
        "own": [item(r) for r in own_source[:2]],
        "competitor": [item(r) for r in competitor_source[:2]],
    }


def _build_email_subject(normalized, logical_date, snapshot=None):
    """Generate a dynamic email subject line with alert level prefix."""
    from qbu_crawler.server.report_common import _compute_alert_level
    alert_level, _ = _compute_alert_level(normalized)
    prefix = {"red": "[需关注] ", "yellow": "[注意] ", "green": ""}[alert_level]
    top = (normalized.get("self", {}).get("risk_products") or [None])[0]
    top_name = top["product_name"] if top else ""
    count = normalized.get("kpis", {}).get("product_count", 0)
    report_window = (snapshot or {}).get("report_window") or {}
    window_type = report_window.get("type")
    title = "产品评论周报" if window_type == "weekly" else ("产品评论监控起点" if window_type == "bootstrap" else "产品评论日报")
    return f"{prefix}{title} {logical_date} — {top_name} 等 {count} 个产品"


def _ensure_humanized_bullets(normalized):
    """Ensure executive_bullets_human exists in normalized analytics."""
    from qbu_crawler.server.report_common import _humanize_bullets
    report_copy = normalized.setdefault("report_copy", {})
    if not report_copy.get("executive_bullets_human"):
        report_copy["executive_bullets_human"] = _humanize_bullets(normalized)


def build_daily_deep_report_email(snapshot, analytics):
    normalized = normalize_deep_report_analytics(analytics)
    _ensure_humanized_bullets(normalized)
    subject = _build_email_subject(normalized, snapshot["logical_date"], snapshot)
    template_env = _report_template_env()
    body = template_env.get_template("daily_report_email_body.txt.j2").render(
        snapshot=snapshot,
        analytics=normalized,
        threshold=config.NEGATIVE_THRESHOLD,
    ).strip()
    return subject, f"{body}\n"


def build_legacy_report_email(
    products: list[dict],
    reviews: list[dict],
    since_str: str,
    translated_count: int,
    untranslated_count: int,
) -> tuple[str, str]:
    """Build the original report email subject/body template."""
    site_names = {"basspro": "Bass Pro Shops", "meatyourmaker": "Meat Your Maker"}
    sites_in_report = set()
    own_count = 0
    competitor_count = 0
    for product in products:
        sites_in_report.add(product.get("site", ""))
        if product.get("ownership") == "own":
            own_count += 1
        else:
            competitor_count += 1

    negative_reviews = [item for item in reviews if (item.get("rating") or 5) <= 2]
    site_display = "、".join(
        site_names.get(site, site) for site in sorted(sites_in_report) if site
    ) or "多站点"

    subject = f"【网评监控】{site_display} 产品评论报告 {since_str}"
    body = (
        f"各位好，\n\n"
        f"附件是 {since_str} 从 {site_display} 抓取的最新产品网评报告，请查阅。\n\n"
        f"【数据概览】\n"
        f"  - 涉及产品：{len(products)} 个"
    )
    if own_count or competitor_count:
        body += f"（自有 {own_count}，竞品 {competitor_count}）"
    body += (
        f"\n"
        f"  - 新增评论：{len(reviews)} 条（已翻译 {translated_count} 条）\n"
    )
    if untranslated_count > 0:
        body += f"  - 注：{untranslated_count} 条评论翻译进行中，中文列暂时为空\n"

    if negative_reviews:
        body += (
            f"\n"
            f"【差评预警】共 {len(negative_reviews)} 条差评（≤2星），请重点关注并更新改进措施。\n"
        )

    body += (
        f"\n"
        f"详细数据见附件 Excel（产品 + 评论两个 Sheet）。\n"
        f"如有疑问请随时沟通，谢谢！\n"
    )
    return subject, body


# ── Pipeline ─────────────────────────────────────────────────────────────────

# Alias before defining generate_report so the parameter name doesn't shadow it
_send_email_impl = send_email


def generate_report(
    since: datetime,
    send_email: bool = True,
) -> dict:
    """Report pipeline: query (with pre-translated data) → Excel → email.

    Translation is handled by the background TranslationWorker.
    Reviews that haven't been translated yet will have empty Chinese fields.
    """
    # 1. Query (includes headline_cn/body_cn from DB)
    products, reviews = query_report_data(since)

    # 2. Count translation status
    translated_count = sum(
        1 for r in reviews if r.get("translate_status") == "done"
    )
    untranslated_count = len(reviews) - translated_count

    # Ensure Chinese fields exist for Excel generation
    for r in reviews:
        r.setdefault("headline_cn", "")
        r.setdefault("body_cn", "")

    # 3. Excel
    excel_path = generate_excel(products, reviews, report_date=since)

    # 4. Email
    email_result = None
    if send_email:
        recipients = config.EMAIL_RECIPIENTS
        since_str = since.strftime("%Y-%m-%d")
        subject, body = build_legacy_report_email(
            products=products,
            reviews=reviews,
            since_str=since_str,
            translated_count=translated_count,
            untranslated_count=untranslated_count,
        )
        email_result = _send_email_impl(
            recipients=recipients,
            subject=subject,
            body_text=body,
            attachment_path=excel_path,
        )

    return {
        "products_count": len(products),
        "reviews_count": len(reviews),
        "translated_count": translated_count,
        "untranslated_count": untranslated_count,
        "excel_path": excel_path,
        "email": email_result,
    }


_legacy_query_report_data = query_report_data
_legacy_generate_excel = generate_excel


# ── 4-Sheet Data-Oriented Excel ──────────────────────────────────────────────


def _generate_analytical_excel(
    products: list[dict],
    reviews: list[dict],
    analytics: dict | None,
    report_date: datetime | None = None,
) -> str:
    """F011 §4.3 — generate the 4-sheet analytical Excel workbook.

    Sheets (in order):
      1. 核心数据   — per-product summary with status lamp + dual-denominator
                      negative rate (H10) + 主要问题 from product_status
      2. 现在该做什么 — report_user_contract.action_priorities with short_title +
                      affected_products + full_action + evidence_count
      3. 评论原文   — raw reviews with normalized impact_category enum (H12) and
                      failure_mode 9-class enum (H19), distinct from labels
      4. 竞品启示   — competitor benchmark_examples (positive) +
                      negative_opportunities (negative) themes
    """
    if analytics is None:
        return _legacy_generate_excel(products, reviews, report_date=report_date)
    analytics = normalize_deep_report_analytics(analytics)

    if report_date is None:
        report_date = config.now_shanghai()
    if report_date.tzinfo is not None:
        report_date = report_date.astimezone(config.SHANGHAI_TZ).replace(tzinfo=None)

    os.makedirs(config.REPORT_DIR, exist_ok=True)
    filename = f"scrape-report-{report_date.strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(config.REPORT_DIR, filename)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    report_semantics = analytics.get("report_semantics") or (
        "bootstrap" if analytics.get("mode", "baseline") == "baseline" else "incremental"
    )
    window_review_ids = set(analytics.get("window_review_ids") or [])
    fresh_cutoff = report_date - timedelta(days=30)

    def _write_headers(ws, headers):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _auto_widths(ws, min_w=10, max_w=60):
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)

    def _parse_json_list(value):
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except Exception:
                return []
        if isinstance(value, list):
            return value
        return []

    def _safe_text(value):
        if value in (None, "", "None", "null"):
            return ""
        return str(value)

    def _parse_review_datetime(review):
        raw = review.get("date_published_parsed") or review.get("date_published")
        if not raw:
            return None
        text = str(raw).strip()
        if not text:
            return None
        if text.endswith("Z"):
            text = f"{text[:-1]}+00:00"
        candidates = [text]
        if len(text) >= 19:
            candidates.append(text[:19])
        if len(text) >= 10:
            candidates.append(text[:10])
        for candidate in candidates:
            try:
                parsed = datetime.fromisoformat(candidate)
                if parsed.tzinfo is not None:
                    parsed = parsed.astimezone(config.SHANGHAI_TZ).replace(tzinfo=None)
                return parsed
            except ValueError:
                continue
        return None

    def _review_new_flag(review):
        in_window = review.get("id") in window_review_ids
        if report_semantics == "bootstrap":
            published = _parse_review_datetime(review)
            window_label = "本次入库" if in_window else "历史累计"
            freshness_label = "新近" if published and published >= fresh_cutoff else "补采"
            return f"{window_label}·{freshness_label}"
        return "本次入库" if in_window else "历史累计"

    def _labels_text(review):
        return ", ".join(
            _xl_display(label.get("code"), _LABEL_DISPLAY)
            for label in _parse_json_list(review.get("analysis_labels") or "[]")
            if isinstance(label, dict)
        )

    def _features_text(review):
        return ", ".join(
            _safe_text(item)
            for item in _parse_json_list(review.get("analysis_features") or review.get("features") or "[]")
            if _safe_text(item)
        )

    # ── Pre-aggregate review counts by SKU (used in 核心数据) ─────────────
    review_counts_by_sku: dict[str, int] = {}
    negative_counts_by_sku: dict[str, int] = {}
    for review in reviews:
        sku = review.get("product_sku") or ""
        if not sku:
            continue
        review_counts_by_sku[sku] = review_counts_by_sku.get(sku, 0) + 1
        if (review.get("rating") or 5) <= config.NEGATIVE_THRESHOLD:
            negative_counts_by_sku[sku] = negative_counts_by_sku.get(sku, 0) + 1

    self_block = analytics.get("self") or {}
    product_status_by_sku = {
        item.get("product_sku"): item
        for item in (self_block.get("product_status") or [])
        if item.get("product_sku")
    }

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 1: 核心数据 (F011 §4.3.2 — 14 cols)
    # ─────────────────────────────────────────────────────────────────────
    ws_core = wb.active
    ws_core.title = "核心数据"
    # 列设计：业务侧（站点评分/总评论数/差评数/差评率_站点/样本覆盖率）+ 运维侧（仅评分数/文字评论数/采集完整率）
    # 命名严格区分，避免读者把"覆盖率"歧义化
    core_headers = [
        "产品名称", "SKU", "站点", "归属", "售价", "库存状态",
        "站点评分", "站点评论数", "仅评分数", "文字评论数",
        "采集评论数", "覆盖率", "采集完整率",
        "差评数", "差评率(站点分母)", "差评率(采集分母)",
        "状态灯", "主要问题",
    ]
    _write_headers(ws_core, core_headers)
    from qbu_crawler.server.report_common import text_review_count
    for product in products:
        sku = product.get("sku") or ""
        site_count = int(product.get("review_count") or 0)
        ratings_only = int(product.get("ratings_only_count") or 0)
        text_count = text_review_count(product)  # = site_count - ratings_only
        ingested = review_counts_by_sku.get(sku, 0)
        negatives = negative_counts_by_sku.get(sku, 0)
        # 业务口径：文本样本占比 = ingested / 站点总反馈（含 ratings-only），反映分析代表性
        text_sample_share = (ingested / site_count) if site_count > 0 else None
        # 运维口径：采集完整率 = ingested / 文字评论数，反映 scraper 健康，应 ≈ 100%
        scrape_completeness = (ingested / text_count) if text_count > 0 else None
        # 差评率分母维持原义（站点 / 采集），不引入 ratings-only 修正以免破坏既有横向对比
        rate_site = (negatives / site_count) if site_count > 0 else None
        rate_ingested = (negatives / ingested) if ingested > 0 else None

        ownership = product.get("ownership") or "own"
        status_entry = product_status_by_sku.get(sku) if ownership == "own" else None
        if status_entry:
            lamp_value = _STATUS_LAMP_DISPLAY.get(
                status_entry.get("status_lamp"),
                status_entry.get("status_label") or "",
            )
            primary_concern = status_entry.get("primary_concern") or ""
        else:
            lamp_value = "" if ownership == "own" else "—"
            primary_concern = ""

        ws_core.append([
            product.get("name"),
            sku,
            product.get("site"),
            _xl_display(ownership, _OWNERSHIP_DISPLAY),
            product.get("price"),
            _xl_display(product.get("stock_status"), _STOCK_DISPLAY),
            product.get("rating"),
            site_count,
            ratings_only,
            text_count,
            ingested,
            round(text_sample_share, 4) if text_sample_share is not None else "—",
            round(scrape_completeness, 4) if scrape_completeness is not None else "—",
            negatives,
            round(rate_site, 4) if rate_site is not None else "—",
            round(rate_ingested, 4) if rate_ingested is not None else "—",
            lamp_value,
            primary_concern,
        ])
    _auto_widths(ws_core)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 2: 现在该做什么 (F011 §4.3.2 — 7 cols, top-5 priorities)
    # ─────────────────────────────────────────────────────────────────────
    ws_reco = wb.create_sheet("现在该做什么")
    reco_headers = [
        "序号", "短标题", "影响产品数", "影响产品列表",
        "用户原话(典型)", "改良方向", "证据数",
    ]
    _write_headers(ws_reco, reco_headers)
    priorities = (
        ((analytics.get("report_user_contract") or {}).get("action_priorities"))
        or []
    )
    for idx, rec in enumerate(priorities[:5], start=1):
        affected = rec.get("affected_products") or []
        ws_reco.append([
            idx,
            _safe_text(rec.get("short_title")),
            int(rec.get("affected_products_count") or len(affected) or 0),
            "、".join(str(p) for p in affected if p),
            _safe_text(rec.get("top_complaint")),
            _safe_text(rec.get("full_action")),
            int(rec.get("evidence_count") or 0),
        ])
    _auto_widths(ws_reco)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 3: 评论原文 (F011 §4.3.2 — 18 cols, retains image embed)
    # ─────────────────────────────────────────────────────────────────────
    ws_rev = wb.create_sheet("评论原文")
    review_headers = [
        "ID", "窗口归属", "产品名称", "SKU", "归属", "评分", "情感", "标签",
        "影响类别", "失效模式",
        "标题(原文)", "标题(中文)", "内容(原文)", "内容(中文)",
        "特征短语", "洞察", "评论时间", "照片",
    ]
    _write_headers(ws_rev, review_headers)
    images_col = len(review_headers)

    all_image_urls: set[str] = set()
    for review in reviews:
        for url in _parse_json_list(review.get("images") or []):
            if isinstance(url, str) and url.startswith("http"):
                all_image_urls.add(url)
    prefetched = _download_images_parallel(list(all_image_urls)) if all_image_urls else {}

    for review in reviews:
        images_list = _parse_json_list(review.get("images") or [])
        labels_text = _labels_text(review)
        features_text = _features_text(review)

        impact_raw = _safe_text(review.get("impact_category"))
        impact_display = IMPACT_CATEGORY_DISPLAY.get(impact_raw, impact_raw)

        failure_raw = _safe_text(review.get("failure_mode"))
        failure_display = FAILURE_MODE_DISPLAY.get(failure_raw, failure_raw)

        row = [
            review.get("id"),
            _review_new_flag(review),
            review.get("product_name"),
            review.get("product_sku"),
            _xl_display(review.get("ownership"), _OWNERSHIP_DISPLAY),
            review.get("rating"),
            _xl_display(review.get("sentiment"), _SENTIMENT_DISPLAY),
            labels_text,
            impact_display,
            failure_display,
            _safe_text(review.get("headline")),
            _safe_text(review.get("headline_cn")),
            _safe_text(review.get("body")),
            _safe_text(review.get("body_cn")),
            features_text,
            _safe_text(review.get("analysis_insight_cn") or review.get("insight_cn")),
            _safe_text(review.get("date_published_parsed") or review.get("date_published")),
            "",
        ]
        ws_rev.append(row)

        row_idx = ws_rev.max_row
        embedded_count = 0
        for url in images_list:
            if not isinstance(url, str) or not url.startswith("http"):
                continue
            cached = prefetched.get(url, _IMG_FAILED)
            if cached is _IMG_FAILED:
                cached = _download_image_data(url)
            if not cached or cached is _IMG_FAILED:
                continue
            xl_img = _make_xl_image(cached)
            y_offset = embedded_count * (_IMG_THUMB_HEIGHT + _IMG_THUMB_SPACING)
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            marker = AnchorMarker(
                col=images_col - 1,
                row=row_idx - 1,
                colOff=0,
                rowOff=pixels_to_EMU(y_offset),
            )
            size = XDRPositiveSize2D(
                pixels_to_EMU(xl_img.width),
                pixels_to_EMU(xl_img.height),
            )
            xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws_rev.add_image(xl_img)
            embedded_count += 1

        if embedded_count > 0:
            row_height_px = embedded_count * (_IMG_THUMB_HEIGHT + _IMG_THUMB_SPACING)
            ws_rev.row_dimensions[row_idx].height = row_height_px * 0.75
        elif images_list:
            ws_rev.cell(
                row=row_idx, column=images_col,
                value="\n".join(str(item) for item in images_list),
            )

    _auto_widths(ws_rev)
    ws_rev.column_dimensions[get_column_letter(images_col)].width = _IMG_COL_WIDTH

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 4: 竞品启示 (F011 §4.3.2 — 5 cols, top-3 + top-3)
    # ─────────────────────────────────────────────────────────────────────
    ws_comp = wb.create_sheet("竞品启示")
    comp_headers = [
        "类型", "主题", "证据数", "典型评论(中文)", "涉及产品",
        "对自有产品启发", "验证动作", "证据ID", "样本数", "涉及产品数",
    ]
    _write_headers(ws_comp, comp_headers)

    competitor = analytics.get("competitor") or {}
    contract_competitor = (analytics.get("report_user_contract") or {}).get("competitor_insights") or {}

    def _theme_topic(item):
        # benchmark_examples / negative_opportunities expose label_codes
        codes = item.get("label_codes") or []
        if codes:
            return _xl_display(codes[0], _LABEL_DISPLAY)
        return _safe_text(item.get("label_display") or item.get("topic"))

    def _theme_example(item):
        return (
            _safe_text(item.get("body_cn"))
            or _safe_text(item.get("headline_cn"))
            or _safe_text(item.get("body"))
            or _safe_text(item.get("headline"))
        )

    def _theme_product(item):
        return _safe_text(item.get("product_name") or item.get("product_sku"))

    contract_rows = []
    benchmark_items = []
    negative_items = []
    for section, label in (
        ("learn_from_competitors", "可借鉴"),
        ("avoid_competitor_failures", "短板"),
        ("validation_hypotheses", "验证假设"),
    ):
        for item in (contract_competitor.get(section) or [])[:3]:
            row = dict(item)
            row["_excel_type"] = label
            contract_rows.append(row)

    if contract_rows:
        for item in contract_rows:
            products_text = "、".join(
                str(pid) for pid in (item.get("products") or item.get("affected_products") or []) if pid
            )
            evidence_ids_text = "、".join(str(rid) for rid in (item.get("evidence_review_ids") or []) if rid)
            ws_comp.append([
                item.get("_excel_type") or "—",
                _safe_text(item.get("theme") or item.get("label_display") or item.get("label_code")) or "—",
                int(item.get("evidence_count") or len(item.get("evidence_review_ids") or []) or 0),
                _safe_text(item.get("competitor_signal") or item.get("summary_cn")) or "—",
                products_text or "—",
                _safe_text(item.get("self_product_implication")) or "—",
                _safe_text(item.get("validation_hypothesis") or item.get("suggested_validation")) or "—",
                evidence_ids_text or "—",
                int(item.get("sample_size") or 0),
                int(item.get("product_count") or len(item.get("products") or []) or 0),
            ])
    else:
        raw_benchmark_items = competitor.get("benchmark_examples") or []
        if isinstance(raw_benchmark_items, dict):
            for category, label in (
                ("product_design", "可借鉴·产品形态"),
                ("marketing_message", "可借鉴·营销话术"),
                ("service_model", "可借鉴·服务模式"),
            ):
                for item in (raw_benchmark_items.get(category) or [])[:1]:
                    row = dict(item)
                    row["_excel_type"] = label
                    benchmark_items.append(row)
        else:
            benchmark_items = raw_benchmark_items[:3]
        negative_items = (competitor.get("negative_opportunities") or [])[:3]

        for item in benchmark_items:
            codes = item.get("label_codes") or []
            ws_comp.append([
                item.get("_excel_type") or "可借鉴",
                _theme_topic(item) or "—",
                len(codes) if codes else 1,
                _theme_example(item) or "—",
                _theme_product(item) or "—",
                "—",
                "—",
                "—",
                len(codes) if codes else 1,
                1 if _theme_product(item) else 0,
            ])

        for item in negative_items:
            codes = item.get("label_codes") or []
            ws_comp.append([
                "短板",
                _theme_topic(item) or "—",
                len(codes) if codes else 1,
                _theme_example(item) or "—",
                _theme_product(item) or "—",
                "—",
                "—",
                "—",
                len(codes) if codes else 1,
                1 if _theme_product(item) else 0,
            ])

    if not contract_rows and not benchmark_items and not negative_items:
        ws_comp.append(["—", "—", 0, "无竞品评论数据", "—", "—", "—", "—", 0, 0])

    _auto_widths(ws_comp)

    wb.save(filepath)
    logger.info("F011 4-sheet Excel generated: %s", filepath)
    return filepath


def _report_ts(value: datetime | str) -> str:
    """Normalize report cutoffs to the naive timestamp format stored in SQLite."""
    if isinstance(value, str):
        value = datetime.fromisoformat(value)
    if value.tzinfo is not None:
        value = value.astimezone(config.SHANGHAI_TZ).replace(tzinfo=None)
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _product_select_columns(conn, alias=None):
    prefix = f"{alias}." if alias else ""
    columns = {row[1] for row in conn.execute("PRAGMA table_info(products)").fetchall()}
    ratings_only = f"{prefix}ratings_only_count" if "ratings_only_count" in columns else "0 AS ratings_only_count"
    return (
        f"{prefix}url, {prefix}name, {prefix}sku, {prefix}price, {prefix}stock_status, "
        f"{prefix}rating, {prefix}review_count, {ratings_only}, "
        f"{prefix}scraped_at, {prefix}site, {prefix}ownership"
    )


def query_report_data(
    since: datetime | str,
    until: datetime | str | None = None,
) -> tuple[list[dict], list[dict]]:
    """Query products and reviews inside a bounded report window."""
    if until is None:
        return _legacy_query_report_data(since if isinstance(since, datetime) else datetime.fromisoformat(since))

    since_str = _report_ts(since)
    until_str = _report_ts(until)

    conn = models.get_conn()
    try:
        product_columns = _product_select_columns(conn)
        product_rows = conn.execute(
            f"""
            SELECT {product_columns}
            FROM products
            WHERE scraped_at >= ?
              AND scraped_at < ?
            ORDER BY scraped_at DESC
            """,
            (since_str, until_str),
        ).fetchall()
        products = [dict(r) for r in product_rows]

        review_rows = conn.execute(
            """
            SELECT r.id AS id, p.name AS product_name, p.sku AS product_sku,
                   p.url AS product_url, p.site AS site, r.scraped_at AS scraped_at,
                   r.author, r.headline, r.body, r.rating,
                   r.date_published, r.date_published_parsed, r.images,
                   p.ownership,
                   r.headline_cn, r.body_cn, r.translate_status,
                   ra.sentiment,
                   ra.sentiment_score,
                   ra.labels   AS analysis_labels,
                   ra.features AS analysis_features,
                   ra.insight_cn AS analysis_insight_cn,
                   ra.insight_en AS analysis_insight_en,
                   ra.impact_category,
                   ra.failure_mode
            FROM reviews r
            JOIN products p ON r.product_id = p.id
            LEFT JOIN review_analysis ra
                ON ra.review_id = r.id
                AND ra.analyzed_at = (
                    SELECT MAX(ra2.analyzed_at)
                    FROM review_analysis ra2
                    WHERE ra2.review_id = r.id
                )
            WHERE r.scraped_at >= ?
              AND r.scraped_at < ?
            ORDER BY r.scraped_at DESC
            """,
            (since_str, until_str),
        ).fetchall()
        reviews = []
        for row in review_rows:
            data = dict(row)
            if data.get("images") and isinstance(data["images"], str):
                try:
                    data["images"] = json.loads(data["images"])
                except Exception:
                    pass
            reviews.append(data)
    finally:
        conn.close()

    logger.info(
        "query_report_data: since=%s until=%s -> %d products, %d reviews",
        since_str,
        until_str,
        len(products),
        len(reviews),
    )
    return products, reviews


def query_cumulative_data() -> tuple[list[dict], list[dict]]:
    """Query all products and all reviews (no time-window filter).

    Returns (products, reviews) — both as lists of dicts.
    Reviews include the latest analysis fields (sentiment, analysis_labels, etc.)
    via a LEFT JOIN with MAX(analyzed_at) subquery to prevent duplicate rows.
    """
    conn = models.get_conn()
    try:
        product_columns = _product_select_columns(conn)
        product_rows = conn.execute(
            f"""
            SELECT {product_columns}
            FROM products
            ORDER BY scraped_at DESC
            """
        ).fetchall()
        products = [dict(r) for r in product_rows]

        review_rows = conn.execute(
            """
            SELECT r.id AS id, p.name AS product_name, p.sku AS product_sku,
                   p.url AS product_url, p.site AS site, r.scraped_at AS scraped_at,
                   r.author, r.headline, r.body, r.rating,
                   r.date_published, r.date_published_parsed, r.images,
                   p.ownership,
                   r.headline_cn, r.body_cn, r.translate_status,
                   ra.sentiment,
                   ra.sentiment_score,
                   ra.labels   AS analysis_labels,
                   ra.features AS analysis_features,
                   ra.insight_cn AS analysis_insight_cn,
                   ra.insight_en AS analysis_insight_en,
                   ra.impact_category,
                   ra.failure_mode
            FROM reviews r
            JOIN products p ON r.product_id = p.id
            LEFT JOIN review_analysis ra
                ON ra.review_id = r.id
                AND ra.analyzed_at = (
                    SELECT MAX(ra2.analyzed_at)
                    FROM review_analysis ra2
                    WHERE ra2.review_id = r.id
                )
            ORDER BY r.scraped_at DESC
            """
        ).fetchall()
        reviews = []
        for row in review_rows:
            data = dict(row)
            if data.get("images") and isinstance(data["images"], str):
                try:
                    data["images"] = json.loads(data["images"])
                except Exception:
                    pass
            reviews.append(data)
    finally:
        conn.close()

    logger.info(
        "query_cumulative_data: %d products, %d reviews",
        len(products),
        len(reviews),
    )
    return products, reviews


def _query_reviews_with_latest_analysis_for_trend(conn, until_str, lookback_days):
    review_rows = conn.execute(
        """
        SELECT r.id AS id, p.name AS product_name, p.sku AS product_sku,
               p.url AS product_url, p.site AS site, p.ownership,
               r.scraped_at AS scraped_at,
               r.author, r.headline, r.body, r.rating,
               r.date_published, r.date_published_parsed, r.images,
               r.headline_cn, r.body_cn, r.translate_status,
               ra.sentiment,
               ra.sentiment_score,
               ra.labels AS analysis_labels,
               ra.features AS analysis_features,
               ra.insight_cn AS analysis_insight_cn,
               ra.insight_en AS analysis_insight_en,
               ra.impact_category,
               ra.failure_mode
        FROM reviews r
        JOIN products p ON r.product_id = p.id
        LEFT JOIN review_analysis ra
            ON ra.review_id = r.id
            AND ra.analyzed_at = (
                SELECT MAX(ra2.analyzed_at)
                FROM review_analysis ra2
                WHERE ra2.review_id = r.id
            )
        WHERE r.scraped_at < ?
          AND datetime(COALESCE(
                CASE
                    WHEN date(r.date_published_parsed) IS NOT NULL
                    THEN r.date_published_parsed
                END,
                CASE
                    WHEN date(substr(r.date_published, 1, 10)) IS NOT NULL
                    THEN substr(r.date_published, 1, 10)
                END,
                substr(r.scraped_at, 1, 10)
              )) < datetime(?)
          AND datetime(COALESCE(
                CASE
                    WHEN date(r.date_published_parsed) IS NOT NULL
                    THEN r.date_published_parsed
                END,
                CASE
                    WHEN date(substr(r.date_published, 1, 10)) IS NOT NULL
                    THEN substr(r.date_published, 1, 10)
                END,
                substr(r.scraped_at, 1, 10)
              )) >= datetime(?, ?)
        ORDER BY COALESCE(
            CASE
                WHEN date(r.date_published_parsed) IS NOT NULL
                THEN r.date_published_parsed
            END,
            CASE
                WHEN date(substr(r.date_published, 1, 10)) IS NOT NULL
                THEN substr(r.date_published, 1, 10)
            END,
            substr(r.scraped_at, 1, 10)
        ) ASC, r.id ASC
        """,
        (until_str, until_str, until_str, f"-{lookback_days} days"),
    ).fetchall()
    reviews = []
    for row in review_rows:
        data = dict(row)
        if data.get("images") and isinstance(data["images"], str):
            try:
                data["images"] = json.loads(data["images"])
            except Exception:
                pass
        reviews.append(data)
    return reviews


def query_trend_history(until, lookback_days=730):
    until_str = _report_ts(until)
    conn = models.get_conn()
    try:
        products = [dict(row) for row in conn.execute(
            """
            SELECT DISTINCT
                   p.url, p.name, p.sku, p.site, p.ownership
            FROM products p
            JOIN product_snapshots ps ON ps.product_id = p.id
            WHERE ps.scraped_at < ?
            ORDER BY p.site, p.ownership, p.sku
            """,
            (until_str,),
        ).fetchall()]
        reviews = _query_reviews_with_latest_analysis_for_trend(conn, until_str, lookback_days)
        return products, reviews
    finally:
        conn.close()


def query_scope_report_data(scope: Scope) -> tuple[list[dict], list[dict]]:
    """Query products and reviews for a normalized scope."""
    conn = models.get_conn()
    try:
        review_clauses, review_params = models._scope_review_clauses(  # noqa: SLF001
            scope,
            review_alias="r",
            product_alias="p",
        )
        review_where = f"WHERE {' AND '.join(review_clauses)}" if review_clauses else ""

        if models._scope_has_review_constraints(scope):  # noqa: SLF001
            product_columns = _product_select_columns(conn, alias="p")
            product_rows = conn.execute(
                f"""
                SELECT DISTINCT
                       {product_columns}
                FROM reviews r
                JOIN products p ON r.product_id = p.id
                {review_where}
                ORDER BY p.scraped_at DESC, p.id DESC
                """,
                review_params,
            ).fetchall()
        else:
            product_clauses, product_params = models._scope_product_clauses(scope, alias="p")  # noqa: SLF001
            product_where = f"WHERE {' AND '.join(product_clauses)}" if product_clauses else ""
            product_columns = _product_select_columns(conn)
            product_rows = conn.execute(
                f"""
                SELECT {product_columns}
                FROM products p
                {product_where}
                ORDER BY p.scraped_at DESC, p.id DESC
                """,
                product_params,
            ).fetchall()
        products = [dict(r) for r in product_rows]

        review_rows = conn.execute(
            f"""
            SELECT p.name AS product_name, p.sku AS product_sku,
                   r.author, r.headline, r.body, r.rating,
                   r.date_published, r.date_published_parsed, r.images,
                   p.ownership,
                   r.headline_cn, r.body_cn, r.translate_status
            FROM reviews r
            JOIN products p ON r.product_id = p.id
            {review_where}
            ORDER BY r.scraped_at DESC, r.id DESC
            """,
            review_params,
        ).fetchall()
        reviews = []
        for row in review_rows:
            data = dict(row)
            if data.get("images") and isinstance(data["images"], str):
                try:
                    data["images"] = json.loads(data["images"])
                except Exception:
                    pass
            reviews.append(data)
    finally:
        conn.close()

    logger.info(
        "query_scope_report_data: %d products, %d reviews",
        len(products),
        len(reviews),
    )
    return products, reviews


def _filtered_report_date(scope: Scope) -> datetime:
    if scope.window.since:
        return datetime.fromisoformat(scope.window.since)
    if scope.window.until:
        return datetime.fromisoformat(scope.window.until)
    return config.now_shanghai()


def _validate_scope_window(scope: Scope) -> str | None:
    parsed = {}
    for label, value in (("since", scope.window.since), ("until", scope.window.until)):
        if not value:
            continue
        try:
            parsed[label] = datetime.fromisoformat(value)
        except ValueError:
            return f"Invalid scope window: {label} must be ISO date or datetime"
    if parsed.get("since") and parsed.get("until") and parsed["since"] > parsed["until"]:
        return "Invalid scope window: since must be on or before until"
    return None


def send_filtered_report(scope: dict, delivery: dict | None = None) -> dict:
    """Generate a filtered report while preserving the legacy email contract."""
    delivery = delivery or {}
    scope_obj = normalize_scope(
        products=scope.get("products"),
        reviews=scope.get("reviews"),
        window=scope.get("window"),
    )
    output_format = str(delivery.get("format", "excel") or "excel").strip().lower()
    subject_override = str(delivery.get("subject", "") or "").strip()
    if "recipients" in delivery:
        recipients = delivery.get("recipients")
    else:
        recipients = config.EMAIL_RECIPIENTS

    window_error = _validate_scope_window(scope_obj)
    if window_error:
        return {
            "scope": asdict(scope_obj),
            "data": {
                "products_count": 0,
                "reviews_count": 0,
                "translated_count": 0,
                "untranslated_count": 0,
            },
            "artifact": {
                "success": False,
                "format": "excel",
                "excel_path": None,
                "error": window_error,
            },
            "email": None,
        }

    products, reviews = query_scope_report_data(scope_obj)

    translated_count = sum(1 for r in reviews if r.get("translate_status") == "done")
    untranslated_count = len(reviews) - translated_count
    for review in reviews:
        review.setdefault("headline_cn", "")
        review.setdefault("body_cn", "")

    data_result = {
        "products_count": len(products),
        "reviews_count": len(reviews),
        "translated_count": translated_count,
        "untranslated_count": untranslated_count,
    }

    report_date = _filtered_report_date(scope_obj)
    output_path = delivery.get("output_path")
    artifact_result = {
        "success": False,
        "format": "excel",
        "excel_path": None,
    }
    try:
        excel_path = generate_excel(
            products,
            reviews,
            report_date=report_date,
            output_path=output_path,
        )
        artifact_result = {
            "success": True,
            "format": "excel",
            "excel_path": excel_path,
        }
    except Exception as exc:
        artifact_result["error"] = str(exc)
        return {
            "scope": asdict(scope_obj),
            "data": data_result,
            "artifact": artifact_result,
            "email": None,
        }

    email_result = None
    if output_format == "email":
        since_str = report_date.strftime("%Y-%m-%d")
        subject, body = build_legacy_report_email(
            products=products,
            reviews=reviews,
            since_str=since_str,
            translated_count=translated_count,
            untranslated_count=untranslated_count,
        )
        if subject_override:
            subject = subject_override
        email_result = _send_email_impl(
            recipients=recipients,
            subject=subject,
            body_text=body,
            attachment_path=artifact_result["excel_path"],
        )

    return {
        "scope": asdict(scope_obj),
        "data": data_result,
        "artifact": artifact_result,
        "email": email_result,
    }


def generate_excel(
    products: list[dict],
    reviews: list[dict],
    report_date: datetime | None = None,
    output_path: str | None = None,
    analytics: dict | None = None,
) -> str:
    """Generate an Excel report, optionally to an immutable output path.

    When *analytics* is provided, produces the 6-sheet analytical workbook;
    otherwise falls back to the legacy 2-sheet format.
    """
    if analytics:
        generated_path = _generate_analytical_excel(
            products, reviews, analytics=analytics, report_date=report_date,
        )
    else:
        generated_path = _legacy_generate_excel(products, reviews, report_date=report_date)

    if not output_path:
        return generated_path

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    if os.path.abspath(generated_path) != os.path.abspath(output_path):
        Path(generated_path).replace(output_path)
    return output_path
