"""Write per-scenario debug/ artifacts."""
import hashlib
import json
import re
import shutil
from pathlib import Path
from .env_bootstrap import load_business
from .db import open_db, row_counts
from . import config


def _json_dump(obj, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(obj, default=str, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def dump_db_state(kind: str, scenario_dir: Path):
    """kind = 'before' | 'after'"""
    with open_db(config.SIM_DB) as conn:
        counts = row_counts(conn)
        samples = {
            "products_by_site_ownership": [
                dict(r) for r in conn.execute(
                    "SELECT site, ownership, COUNT(*) n FROM products GROUP BY site, ownership"
                ).fetchall()
            ],
            "reviews_by_scraped_date": [
                dict(r) for r in conn.execute(
                    "SELECT substr(scraped_at,1,10) d, COUNT(*) n "
                    "FROM reviews GROUP BY d ORDER BY d DESC LIMIT 10"
                ).fetchall()
            ],
            "labels_top": [
                dict(r) for r in conn.execute(
                    "SELECT label_code, label_polarity, COUNT(*) n "
                    "FROM review_issue_labels GROUP BY label_code, label_polarity "
                    "ORDER BY n DESC LIMIT 10"
                ).fetchall()
            ],
        }
    _json_dump({"counts": counts, "samples": samples},
               scenario_dir / "debug" / f"db_state_{kind}.json")


def dump_workflow_run(run_id: int, scenario_dir: Path):
    biz = load_business()
    with biz.models.get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM workflow_runs WHERE id=?", (run_id,)
        ).fetchone()
    _json_dump(dict(row) if row else None,
               scenario_dir / "debug" / "workflow_run.json")


def dump_outbox_rows(run_id: int, scenario_dir: Path):
    biz = load_business()
    with biz.models.get_conn() as conn:
        rows = conn.execute(
            """SELECT id, kind, channel, status, attempts, delivered_at,
                      last_error, created_at, payload
               FROM notification_outbox ORDER BY id"""
        ).fetchall()
    parsed = []
    for r in rows:
        try:
            pl = json.loads(r["payload"] or "{}")
        except ValueError:
            pl = None
        d = dict(r)
        d["payload_parsed"] = pl
        if isinstance(pl, dict) and pl.get("run_id") not in (None, run_id):
            continue
        parsed.append(d)
    _json_dump(parsed, scenario_dir / "debug" / "outbox_rows.json")


def dump_analytics_tree(run_id: int, scenario_dir: Path):
    biz = load_business()
    with biz.models.get_conn() as conn:
        row = conn.execute(
            "SELECT analytics_path FROM workflow_runs WHERE id=?", (run_id,),
        ).fetchone()
    if row and row["analytics_path"] and Path(row["analytics_path"]).exists():
        shutil.copy2(row["analytics_path"],
                     scenario_dir / "debug" / "analytics_tree.json")


def dump_top_reviews(run_id: int, scenario_dir: Path, *, limit: int = 20):
    """Dump the top reviews (by scraped_at desc within window) for this run."""
    biz = load_business()
    with biz.models.get_conn() as conn:
        run = conn.execute(
            "SELECT data_since, data_until FROM workflow_runs WHERE id=?",
            (run_id,),
        ).fetchone()
        if not run or not run["data_since"]:
            _json_dump([], scenario_dir / "debug" / "top_reviews.json")
            return
        rows = conn.execute(
            """SELECT r.id, r.product_id, p.sku, r.rating, r.headline,
                      substr(r.body,1,120) body_preview, r.scraped_at,
                      ra.sentiment, ra.labels, ra.insight_cn
               FROM reviews r
               LEFT JOIN products p ON p.id=r.product_id
               LEFT JOIN review_analysis ra ON ra.review_id=r.id
               WHERE r.scraped_at >= ? AND r.scraped_at < ?
               ORDER BY r.scraped_at DESC LIMIT ?""",
            (run["data_since"], run["data_until"], limit),
        ).fetchall()
    _json_dump([dict(r) for r in rows],
               scenario_dir / "debug" / "top_reviews.json")


def dump_html_checksum(scenario_dir: Path):
    """Structural fingerprint of any .html in scenario root."""
    summary = {}
    for html in scenario_dir.glob("*.html"):
        raw = html.read_text(encoding="utf-8", errors="ignore")
        tags = re.findall(r"<(\w+)", raw)
        tag_counts = {}
        for t in tags:
            tag_counts[t] = tag_counts.get(t, 0) + 1
        summary[html.name] = {
            "byte_len": len(raw),
            "sha1_first32": hashlib.sha1(raw.encode("utf-8")).hexdigest()[:32],
            "tag_counts": dict(sorted(tag_counts.items(), key=lambda kv: -kv[1])[:15]),
        }
    (scenario_dir / "debug").mkdir(exist_ok=True)
    (scenario_dir / "debug" / "html_checksum.txt").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def dump_excel_structure(scenario_dir: Path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    summary = {}
    for xlsx in scenario_dir.glob("*.xlsx"):
        try:
            wb = load_workbook(xlsx, read_only=True)
            summary[xlsx.name] = {
                name: {"max_row": wb[name].max_row, "max_col": wb[name].max_column}
                for name in wb.sheetnames
            }
        except Exception as e:
            summary[xlsx.name] = {"_error": str(e)}
    _json_dump(summary, scenario_dir / "debug" / "excel_structure.json")


def dump_events_applied(events: list, scenario_dir: Path):
    _json_dump(events, scenario_dir / "debug" / "events_applied.json")
