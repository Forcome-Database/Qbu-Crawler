"""End-to-end integration test for Report Intelligence Redesign V2.

Tests the full pipeline: data → analytics → charts → HTML → Excel
without requiring Playwright (PDF) or SMTP (email).
"""
import json
import sqlite3
import pytest
from pathlib import Path
from unittest.mock import patch
from datetime import datetime

from qbu_crawler import models, config
from qbu_crawler.server import report, report_analytics, report_common, report_html
from qbu_crawler.server.report_charts import build_chart_html_fragments
from qbu_crawler.server.report_llm import generate_report_insights


# ── Fixtures ──────────────────────────────────────────────────────────────────


def _get_test_conn(db_file: str):
    conn = sqlite3.connect(db_file)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    return conn


@pytest.fixture
def populated_db(tmp_path, monkeypatch):
    """Create a DB with products, reviews, and review_analysis data."""
    db_path = str(tmp_path / "integration.db")
    monkeypatch.setattr(config, "DB_PATH", db_path)
    monkeypatch.setattr(models, "get_conn", lambda: _get_test_conn(db_path))
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))
    monkeypatch.setattr(config, "REPORT_LABEL_MODE", "rule")
    Path(config.REPORT_DIR).mkdir(parents=True, exist_ok=True)
    models.init_db()

    conn = _get_test_conn(db_path)
    # 3 own products
    conn.execute(
        "INSERT INTO products (id, url, site, name, sku, price, rating, review_count, stock_status, ownership, scraped_at) "
        "VALUES (1, 'http://a', 'basspro', 'Sausage Stuffer', 'BP-SS-01', 169.99, 3.3, 79, 'in_stock', 'own', '2026-04-01 09:00:00')"
    )
    conn.execute(
        "INSERT INTO products (id, url, site, name, sku, price, rating, review_count, stock_status, ownership, scraped_at) "
        "VALUES (2, 'http://b', 'basspro', 'Meat Mixer', 'BP-MM-01', 149.99, 4.3, 57, 'in_stock', 'own', '2026-04-01 09:00:00')"
    )
    conn.execute(
        "INSERT INTO products (id, url, site, name, sku, price, rating, review_count, stock_status, ownership, scraped_at) "
        "VALUES (3, 'http://c', 'basspro', 'Commercial Stuffer', 'BP-CS-01', 349.99, 3.6, 88, 'in_stock', 'own', '2026-04-01 09:00:00')"
    )
    # 2 competitor products
    conn.execute(
        "INSERT INTO products (id, url, site, name, sku, price, rating, review_count, stock_status, ownership, scraped_at) "
        "VALUES (4, 'http://d', 'meatyourmaker', 'Grinder 1HP', 'MYM-1HP', 389.99, 4.7, 683, 'in_stock', 'competitor', '2026-04-01 09:00:00')"
    )
    conn.execute(
        "INSERT INTO products (id, url, site, name, sku, price, rating, review_count, stock_status, ownership, scraped_at) "
        "VALUES (5, 'http://e', 'waltons', 'Patty Maker', 'WAL-PM', 69.99, 4.7, 93, 'in_stock', 'competitor', '2026-04-01 09:00:00')"
    )

    # Reviews: mix of positive and negative
    reviews_data = [
        # Own product negative reviews
        (1, 1, "John",  "Handle broke",       "Handle came off after 3 months",    "abc1", 1.0, "2026-03-15", None, "2026-04-01 09:30:00", "手柄断了",   "用了3个月手柄就断了",    "done"),
        (2, 1, "Mike",  "Terrible quality",   "The handle is loose from day one",  "abc2", 1.0, "2026-03-20", None, "2026-04-01 09:30:00", "质量很差",   "手柄从第一天就松了",    "done"),
        (3, 1, "Jane",  "Leaking",            "Seal leaks during stuffing",        "abc3", 2.0, "2026-02-10", None, "2026-04-01 09:30:00", "漏油",       "灌肉时密封漏油",        "done"),
        (4, 3, "Bob",   "Seal problem",       "Bottom seal leaks oil",             "abc4", 2.0, "2026-01-15", None, "2026-04-01 09:30:00", "密封问题",   "底部密封漏油",          "done"),
        (5, 2, "Alice", "Great mixer",        "Love this mixer, works perfectly",  "abc5", 5.0, "2026-03-01", None, "2026-04-01 09:30:00", "很棒的搅拌机", "非常喜欢这台搅拌机",  "done"),
        # Competitor positive reviews
        (6, 4, "Tom",   "Built like a tank",  "Incredibly solid build quality",   "abc6", 5.0, "2026-03-10", None, "2026-04-01 09:30:00", "做工像坦克",  "做工非常扎实",         "done"),
        (7, 4, "Sam",   "Powerful motor",     "The motor handles everything",      "abc7", 5.0, "2026-02-20", None, "2026-04-01 09:30:00", "电机强劲",   "电机处理能力强",        "done"),
        (8, 5, "Lisa",  "Easy to use",        "Super simple to operate",           "abc8", 5.0, "2026-03-25", None, "2026-04-01 09:30:00", "易于使用",   "操作非常简单",          "done"),
    ]
    for r in reviews_data:
        conn.execute(
            "INSERT INTO reviews (id, product_id, author, headline, body, body_hash, rating, "
            "date_published, images, scraped_at, headline_cn, body_cn, translate_status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            r,
        )

    # review_analysis data (optional enrichment)
    analyses = [
        (1, "negative", 0.10, '[{"code":"quality_stability","polarity":"negative","severity":"high","confidence":0.95}]',
         '["手柄松动","手柄脱落"]', "手柄使用3个月后断裂", "Handle broke after 3 months"),
        (2, "negative", 0.08, '[{"code":"quality_stability","polarity":"negative","severity":"high","confidence":0.92}]',
         '["手柄松动"]', "手柄第一天就松了", "Handle loose from day one"),
        (3, "negative", 0.15, '[{"code":"quality_stability","polarity":"negative","severity":"medium","confidence":0.88}]',
         '["密封漏油"]', "灌肉时密封漏油", "Seal leaks during stuffing"),
        (4, "negative", 0.12, '[{"code":"quality_stability","polarity":"negative","severity":"medium","confidence":0.85}]',
         '["密封漏油"]', "底部密封漏油", "Bottom seal leaks oil"),
        (5, "positive", 0.92, '[{"code":"solid_build","polarity":"positive","severity":"low","confidence":0.90}]',
         '["搅拌性能好"]', "搅拌机工作完美", "Mixer works perfectly"),
        (6, "positive", 0.95, '[{"code":"solid_build","polarity":"positive","severity":"low","confidence":0.95}]',
         '["做工扎实","材质厚实"]', "做工非常扎实", "Incredibly solid build"),
        (7, "positive", 0.93, '[{"code":"strong_performance","polarity":"positive","severity":"low","confidence":0.92}]',
         '["电机强劲"]', "电机处理能力强", "Powerful motor"),
        (8, "positive", 0.91, '[{"code":"easy_to_use","polarity":"positive","severity":"low","confidence":0.88}]',
         '["操作简单"]', "操作非常简单", "Super simple to operate"),
    ]
    for a in analyses:
        conn.execute(
            "INSERT INTO review_analysis "
            "(review_id, sentiment, sentiment_score, labels, features, insight_cn, insight_en, "
            "prompt_version, llm_model, analyzed_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, 'v1', 'gpt-4o-mini', '2026-04-01 10:00:00')",
            a,
        )
    conn.commit()
    conn.close()
    return db_path


@pytest.fixture
def sample_snapshot(populated_db):
    """Build a snapshot dict from the populated DB."""
    conn = _get_test_conn(populated_db)
    products = [dict(r) for r in conn.execute("SELECT * FROM products").fetchall()]
    reviews = [
        dict(r)
        for r in conn.execute(
            "SELECT r.*, p.name AS product_name, p.sku AS product_sku, p.ownership "
            "FROM reviews r JOIN products p ON r.product_id = p.id"
        ).fetchall()
    ]
    conn.close()
    return {
        "run_id": 1,
        "logical_date": "2026-04-01",
        "data_since": "2026-04-01T00:00:00",
        "data_until": "2026-04-02T00:00:00",
        "snapshot_at": "2026-04-01T12:00:00",
        "products": products,
        "reviews": reviews,
        "products_count": len(products),
        "reviews_count": len(reviews),
        "translated_count": len(reviews),
        "untranslated_count": 0,
        "snapshot_hash": "test-hash-123",
    }


# ── Analytics Pipeline ────────────────────────────────────────────────────────


class TestAnalyticsPipeline:
    """Test analytics computation from enriched data."""

    def test_build_report_analytics_basic_kpis(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        assert "self" in analytics
        assert "kpis" in analytics
        assert analytics["kpis"]["product_count"] == 5
        assert analytics["kpis"]["negative_review_rows"] > 0

    def test_build_report_analytics_has_self_section(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        assert "risk_products" in analytics["self"]
        assert "top_negative_clusters" in analytics["self"]

    def test_normalize_analytics_has_health_index(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        assert "health_index" in normalized["kpis"]
        health = normalized["kpis"]["health_index"]
        assert 0 <= health <= 100

    def test_normalize_analytics_has_kpi_cards(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        assert "kpi_cards" in normalized
        assert len(normalized["kpi_cards"]) >= 4

    def test_normalize_analytics_has_issue_cards(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        # issue_cards lives under self
        assert "issue_cards" in normalized.get("self", {})

    # F011 §4.2.5 — retired: legacy 12-panel data shape replaced by primary_chart+drill_downs
    # Original test asserted on analytics["trend_digest"]["data"]["month"]["sentiment"]
    # ["primary_chart"] (legacy shape: labels[] + series[].data[]), and verified
    # that the sentiment series anchors on date_published (excluding the future
    # 2026-04-01 date). The new shape exposes `analytics.trend_digest.primary_chart`
    # with `series_own[].date` (already date-anchored). Date-anchor coverage
    # lives in tests/server/test_date_published_anchor.py and
    # tests/server/test_trend_digest_thresholds.py.
    def test_trend_digest_uses_primary_chart_shape(self, populated_db, sample_snapshot):
        """Smoke: build_report_analytics emits the new trend_digest shape with
        a primary_chart keyed off `kind=health_trend`."""
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        trend = analytics["trend_digest"]
        assert "primary_chart" in trend
        assert "drill_downs" in trend
        assert trend["primary_chart"].get("kind") == "health_trend"

    def test_normalize_analytics_negative_rate_display(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        # 修 8: 顶层混合口径已重命名为 all_sample_negative_rate_display
        assert "all_sample_negative_rate_display" in normalized["kpis"]
        assert "%" in normalized["kpis"]["all_sample_negative_rate_display"]

    def test_normalize_analytics_hero_headline_nonempty(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        hero = normalized.get("report_copy", {}).get("hero_headline", "")
        assert isinstance(hero, str) and len(hero) > 0

    def test_normalize_analytics_display_kpis_do_not_read_cumulative_kpis(self):
        normalized = report_common.normalize_deep_report_analytics(
            {
                "mode": "incremental",
                "kpis": {
                    "product_count": 1,
                    "ingested_review_rows": 12,
                    "translated_count": 12,
                    "own_review_rows": 12,
                    "competitor_review_rows": 0,
                    "own_positive_review_rows": 0,
                    "own_negative_review_rows": 0,
                    "site_reported_review_total_current": 20,
                },
                "cumulative_kpis": {
                    "ingested_review_rows": 999,
                    "own_review_rows": 999,
                },
                "self": {"risk_products": [], "top_negative_clusters": [], "recommendations": []},
                "competitor": {
                    "top_positive_themes": [],
                    "benchmark_examples": [],
                    "negative_opportunities": [],
                    "gap_analysis": [],
                },
                "appendix": {"image_reviews": []},
            }
        )

        assert normalized["kpis"]["ingested_review_rows"] == 12
        assert normalized["cumulative_kpis"]["ingested_review_rows"] == 999
        assert any(card["value"] == 12 for card in normalized["kpi_cards"])
        assert all(card["value"] != 999 for card in normalized["kpi_cards"])


# ── Chart Generation ──────────────────────────────────────────────────────────


class TestChartGeneration:
    """Test Plotly chart generation from analytics."""

    def test_build_chart_fragments_returns_health_gauge(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        charts = build_chart_html_fragments(normalized)
        assert "health_gauge" in charts
        assert "<div" in charts["health_gauge"]

    def test_all_chart_fragments_are_html(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        charts = build_chart_html_fragments(normalized)
        for name, html in charts.items():
            assert "<div" in html, f"Chart {name} is not valid HTML"

    def test_self_risk_products_chart_present(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        charts = build_chart_html_fragments(normalized)
        # Risk products should produce a bar chart when data exists
        if normalized["self"]["risk_products"]:
            assert "self_risk_products" in charts

    def test_negative_clusters_chart_present(self, populated_db, sample_snapshot):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)
        charts = build_chart_html_fragments(normalized)
        if normalized["self"]["top_negative_clusters"]:
            assert "self_negative_clusters" in charts


# ── Report HTML ───────────────────────────────────────────────────────────────


class TestReportHTML:
    """Test V3 HTML rendering via report_html.render_v3_html."""

    def test_render_v3_html_produces_valid_html(self, populated_db, sample_snapshot, tmp_path):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        output_path = str(tmp_path / "test_report.html")

        result = report_html.render_v3_html(sample_snapshot, analytics, output_path=output_path)

        html = Path(result).read_text(encoding="utf-8")
        assert "<!doctype html>" in html.lower() or "<html" in html.lower()

    def test_render_v3_html_contains_chart_js(self, populated_db, sample_snapshot, tmp_path):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        output_path = str(tmp_path / "test_report.html")

        result = report_html.render_v3_html(sample_snapshot, analytics, output_path=output_path)

        html = Path(result).read_text(encoding="utf-8")
        # V3 uses Chart.js instead of Plotly
        assert "chart" in html.lower()

    def test_render_v3_html_contains_hero_headline(self, populated_db, sample_snapshot, tmp_path):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        analytics["report_copy"] = {
            "hero_headline": "测试集成标题",
            "executive_bullets": ["要点1"],
            "executive_summary": "",
            "improvement_priorities": [],
            "competitive_insight": "",
        }
        output_path = str(tmp_path / "test_report.html")

        result = report_html.render_v3_html(sample_snapshot, analytics, output_path=output_path)

        html = Path(result).read_text(encoding="utf-8")
        assert "测试集成标题" in html

    def test_render_v3_html_contains_health_section(self, populated_db, sample_snapshot, tmp_path):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        output_path = str(tmp_path / "test_report.html")

        result = report_html.render_v3_html(sample_snapshot, analytics, output_path=output_path)

        html = Path(result).read_text(encoding="utf-8")
        assert "健康" in html or "health" in html.lower()


# ── Excel Generation ──────────────────────────────────────────────────────────


class TestExcelGeneration:
    """Test 5-sheet data-oriented Excel generation via generate_excel with analytics."""

    def test_generate_analytical_excel_creates_file(self, populated_db, sample_snapshot, tmp_path, monkeypatch):
        monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)

        path = report.generate_excel(
            products=sample_snapshot["products"],
            reviews=sample_snapshot["reviews"],
            analytics=normalized,
        )
        assert Path(path).exists()

    def test_generate_analytical_excel_has_five_sheets(self, populated_db, sample_snapshot, tmp_path, monkeypatch):
        monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)

        path = report.generate_excel(
            products=sample_snapshot["products"],
            reviews=sample_snapshot["reviews"],
            analytics=normalized,
        )

        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 5
        assert set(wb.sheetnames) == {"评论明细", "产品概览", "今日变化", "问题标签", "趋势数据"}

    def test_review_detail_sheet_has_sentiment_column(self, populated_db, sample_snapshot, tmp_path, monkeypatch):
        monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)

        path = report.generate_excel(
            products=sample_snapshot["products"],
            reviews=sample_snapshot["reviews"],
            analytics=normalized,
        )

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["评论明细"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "情感" in headers

    def test_review_detail_sheet_has_data_rows(self, populated_db, sample_snapshot, tmp_path, monkeypatch):
        monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        normalized = report_common.normalize_deep_report_analytics(analytics)

        path = report.generate_excel(
            products=sample_snapshot["products"],
            reviews=sample_snapshot["reviews"],
            analytics=normalized,
        )

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["评论明细"]
        # Header + at least one data row
        assert ws.max_row >= 2
        # No embedded images in the governed analytical format
        assert len(ws._images) == 0


# ── LLM Insights (Fallback) ───────────────────────────────────────────────────


class TestLLMInsights:
    """Test LLM insight generation fallback when LLM is not configured."""

    def test_generate_report_insights_fallback_returns_required_keys(self, populated_db, sample_snapshot, monkeypatch):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        # Disable LLM so fallback path is exercised
        monkeypatch.setattr(config, "LLM_API_BASE", "")
        monkeypatch.setattr(config, "LLM_API_KEY", "")

        insights = generate_report_insights(analytics)

        # Fallback must produce a valid structure
        assert "hero_headline" in insights
        assert "executive_bullets" in insights
        assert isinstance(insights["executive_bullets"], list)

    def test_generate_report_insights_fallback_hero_is_nonempty(self, populated_db, sample_snapshot, monkeypatch):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        monkeypatch.setattr(config, "LLM_API_BASE", "")
        monkeypatch.setattr(config, "LLM_API_KEY", "")

        insights = generate_report_insights(analytics)
        assert len(insights["hero_headline"]) > 0

    def test_generate_report_insights_fallback_bullets_nonempty(self, populated_db, sample_snapshot, monkeypatch):
        analytics = report_analytics.build_report_analytics(sample_snapshot)
        monkeypatch.setattr(config, "LLM_API_BASE", "")
        monkeypatch.setattr(config, "LLM_API_KEY", "")

        insights = generate_report_insights(analytics)
        assert len(insights["executive_bullets"]) > 0
