"""F011 §4.3 — Excel 4 sheets 重构 tests.

Tests for the new Excel layout:
- 核心数据 (core data with status lamp + dual-denominator negative rate)
- 现在该做什么 (recommendations with short_title + full_action + evidence_count)
- 评论原文 (raw reviews with failure_mode enum + impact_category distinct from labels)
- 竞品启示 (competitor benchmarks with type/topic/evidence_count)
"""
import json

import openpyxl
import pytest

from qbu_crawler import config
from qbu_crawler.server.report import generate_excel


_NEW_SHEETS = ["核心数据", "现在该做什么", "评论原文", "竞品启示"]


def _build_analytics(**overrides):
    """Build a minimal analytics dict with the new 4-sheet schema in mind."""
    analytics = {
        "mode": "incremental",
        "report_semantics": "incremental",
        "kpis": {
            "ingested_review_rows": 5,
            "own_review_rows": 5,
            "competitor_review_rows": 0,
            "health_index": 75,
            "high_risk_count": 0,
            "own_product_count": 1,
            "competitor_product_count": 0,
            "translated_count": 5,
            "untranslated_count": 0,
        },
        "self": {
            "risk_products": [],
            "product_status": [
                {
                    "product_name": "ProdA",
                    "product_sku": "SKU-A",
                    "status_lamp": "yellow",
                    "status_label": "需关注",
                    "primary_concern": "外观瑕疵",
                    "risk_score": 18.0,
                    "risk_factors": None,
                    "near_high_risk": False,
                },
            ],
        },
        "competitor": {
            "benchmark_examples": [
                {
                    "product_name": "CompP1",
                    "product_sku": "C-1",
                    "rating": 5,
                    "headline_cn": "做工很扎实",
                    "body_cn": "用了三个月还是很稳",
                    "label_codes": ["build_quality"],
                },
            ],
            "negative_opportunities": [
                {
                    "product_name": "CompP2",
                    "product_sku": "C-2",
                    "rating": 1,
                    "headline_cn": "电机异响",
                    "body_cn": "三天就开始嗡嗡响",
                    "label_codes": ["motor_anomaly"],
                },
            ],
        },
        "report_copy": {
            "improvement_priorities": [
                {
                    "label_code": "cosmetic_finish",
                    "label_display": "外观瑕疵",
                    "short_title": "外观瑕疵：抛光复检",
                    "full_action": (
                        "针对 ProdA 等出现的外观瑕疵问题，建议对供应商抛光工艺做一次"
                        "完整复检；引入 QC 抽检比例提升至 5% 并建立返工闭环。"
                    ),
                    "evidence_count": 4,
                    "evidence_review_ids": [101, 102, 103, 104],
                    "affected_products": ["ProdA"],
                    "affected_products_count": 1,
                    "top_complaint": "外壳上有一道明显划痕",
                },
            ],
        },
        "_trend_series": [],
    }
    analytics.update(overrides)
    return analytics


def _build_products():
    return [
        {
            "name": "ProdA", "sku": "SKU-A", "site": "test", "ownership": "own",
            "price": 99.0, "stock_status": "in_stock", "rating": 4.0,
            "review_count": 100, "scraped_at": "2026-04-25",
        },
        {
            "name": "CompP1", "sku": "C-1", "site": "test", "ownership": "competitor",
            "price": 110.0, "stock_status": "in_stock", "rating": 4.5,
            "review_count": 80, "scraped_at": "2026-04-25",
        },
    ]


def _build_reviews():
    """Build reviews with realistic impact_category + failure_mode enum coverage."""
    labels_cosmetic = json.dumps([
        {"code": "cosmetic_finish", "polarity": "negative", "severity": "medium", "confidence": 0.9},
    ])
    labels_durability = json.dumps([
        {"code": "durability", "polarity": "negative", "severity": "high", "confidence": 0.85},
    ])
    labels_positive = json.dumps([
        {"code": "build_quality", "polarity": "positive", "severity": "low", "confidence": 0.8},
    ])

    base = {
        "product_name": "ProdA", "product_sku": "SKU-A",
        "ownership": "own", "translate_status": "done",
        "analysis_features": "[]", "analysis_insight_cn": "",
    }
    return [
        {
            **base, "id": 101, "author": "u1",
            "headline": "Scratch on shell", "body": "There is a scratch",
            "headline_cn": "外壳划痕", "body_cn": "外壳上有一道划痕",
            "rating": 2.0, "date_published_parsed": "2026-04-20",
            "sentiment": "negative", "images": None,
            "analysis_labels": labels_cosmetic,
            "impact_category": "cosmetic", "failure_mode": "material_finish",
        },
        {
            **base, "id": 102, "author": "u2",
            "headline": "Gear failed", "body": "Gear broke after 2 weeks",
            "headline_cn": "齿轮坏了", "body_cn": "两周就坏了",
            "rating": 1.0, "date_published_parsed": "2026-04-21",
            "sentiment": "negative", "images": None,
            "analysis_labels": labels_durability,
            "impact_category": "functional", "failure_mode": "gear_failure",
        },
        {
            **base, "id": 103, "author": "u3",
            "headline": "Knob broke", "body": "Control knob cracked",
            "headline_cn": "旋钮裂开", "body_cn": "控制旋钮裂了",
            "rating": 2.0, "date_published_parsed": "2026-04-22",
            "sentiment": "negative", "images": None,
            "analysis_labels": labels_cosmetic,
            "impact_category": "durability", "failure_mode": "control_electrical",
        },
        {
            **base, "id": 104, "author": "u4",
            "headline": "Solid build", "body": "Heavy and well-made",
            "headline_cn": "做工扎实", "body_cn": "很重很结实",
            "rating": 5.0, "date_published_parsed": "2026-04-23",
            "sentiment": "positive", "images": None,
            "analysis_labels": labels_positive,
            "impact_category": "service", "failure_mode": "none",
        },
        {
            **base, "id": 105, "author": "u5",
            "headline": "Average", "body": "Just OK",
            "headline_cn": "一般", "body_cn": "普通",
            "rating": 3.0, "date_published_parsed": "2026-04-24",
            "sentiment": "neutral", "images": None,
            "analysis_labels": labels_cosmetic,
            "impact_category": "safety", "failure_mode": "casing_assembly",
        },
    ]


@pytest.fixture
def excel_path(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    products = _build_products()
    reviews = _build_reviews()
    analytics = _build_analytics()
    return generate_excel(
        products, reviews, analytics=analytics,
        output_path=str(tmp_path / "f011-excel.xlsx"),
    )


def test_excel_has_4_sheets(excel_path):
    """F011 §4.3 — Excel 仅 4 sheets。"""
    wb = openpyxl.load_workbook(excel_path)
    assert sorted(wb.sheetnames) == sorted(_NEW_SHEETS)


def test_excel_no_old_sheets(excel_path):
    """已删除：今日变化 / 问题标签 / 趋势数据。"""
    wb = openpyxl.load_workbook(excel_path)
    assert "今日变化" not in wb.sheetnames
    assert "问题标签" not in wb.sheetnames
    assert "趋势数据" not in wb.sheetnames
    assert "评论明细" not in wb.sheetnames
    assert "产品概览" not in wb.sheetnames


def test_core_data_sheet_unified_denominator(excel_path):
    """F011 H10 — 核心数据 sheet 差评率分两列，列名声明分母。"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["核心数据"]
    headers = [c.value for c in ws[1]]
    assert "差评率(采集分母)" in headers
    assert "差评率(站点分母)" in headers
    assert "状态灯" in headers
    assert "覆盖率" in headers


def test_review_original_sheet_failure_mode_filled(excel_path):
    """F011 H12+H19 — 评论原文 sheet 的失效模式列非空率 ≥ 95%."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["评论原文"]
    headers = [c.value for c in ws[1]]
    fm_col = headers.index("失效模式") + 1
    non_empty = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, fm_col).value)
    assert ws.max_row >= 2
    assert non_empty / (ws.max_row - 1) >= 0.95


def test_review_original_sheet_impact_category_distinct_from_labels(excel_path):
    """F011 H12 — 影响类别 不再与 标签 列雷同."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["评论原文"]
    headers = [c.value for c in ws[1]]
    label_col = headers.index("标签") + 1
    impact_col = headers.index("影响类别") + 1
    distinct = sum(
        1 for r in range(2, ws.max_row + 1)
        if ws.cell(r, label_col).value != ws.cell(r, impact_col).value
    )
    assert distinct >= (ws.max_row - 1) * 0.9


def test_recommendations_sheet_uses_short_title(excel_path):
    """F011 §4.3 — 现在该做什么 sheet 列：短标题 + 影响产品 + 全文 action."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["现在该做什么"]
    headers = [c.value for c in ws[1]]
    assert "短标题" in headers
    assert "改良方向" in headers  # full_action
    assert "证据数" in headers
    # First data row reflects the priority above
    short_title_col = headers.index("短标题") + 1
    assert ws.cell(2, short_title_col).value == "外观瑕疵：抛光复检"


def test_competitor_insights_sheet_columns(excel_path):
    """F011 §4.3 — 竞品启示 sheet 含 类型 / 主题 / 证据数 / 典型评论 / 涉及产品."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["竞品启示"]
    headers = [c.value for c in ws[1]]
    assert "类型" in headers
    assert "主题" in headers
    assert "证据数" in headers
    assert "典型评论(中文)" in headers
    assert "涉及产品" in headers
