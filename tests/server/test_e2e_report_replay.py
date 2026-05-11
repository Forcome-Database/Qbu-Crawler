"""F011 §4.3 — End-to-end replay test using a synthetic snapshot fixture.

Verifies that ``freeze_report_snapshot`` + ``generate_full_report_from_snapshot``
produces all expected artifacts (snapshot/analytics/xlsx/html_attachment)
satisfying the major F011 ACs that don't depend on a specific production state.

Scope deviation from original plan
----------------------------------
The original plan called for replaying the production "测试5" SQLite dump
(~1MB, tightly coupled to specific row IDs / dates). We instead seed a
realistic but minimal synthetic DB. This validates that the
freeze→generate→record pipeline wires together correctly without coupling
to a snapshot of production reality.

Out of scope (covered by other tests)
-------------------------------------
* AC-2 / AC-3   (Excel column semantics)              — Task 3.6 unit tests
* AC-26          (state machine)                       — Phase 1 tests
* AC-31          (ops alert in zero-scrape)            — Task 4.1 unit tests
* AC-9           (bootstrap mode "首日基线已建档")     — covered by report_llm tests;
                  exercising it here would require staging "first run"
                  state which makes the fixture brittle.
"""
from __future__ import annotations

import json
import os
import sqlite3
import time
from pathlib import Path

import openpyxl
import pytest

from qbu_crawler import config, models
from qbu_crawler.server.report_artifacts import list_artifacts


# ── Test fixture ─────────────────────────────────────────────────────────────


def _get_test_conn(db_file: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_file)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    return conn


@pytest.fixture
def e2e_db(tmp_path, monkeypatch):
    """Create a minimal but realistic test DB seeded with products + reviews
    + review_analysis rows, plus a workflow_run row covering a 24h window
    around 2026-04-27.

    Returns dict with: db_file, run_id, tmp_path.
    """
    db_file = str(tmp_path / "products.db")
    monkeypatch.setattr(config, "DB_PATH", db_file)
    monkeypatch.setattr(models, "get_conn", lambda: _get_test_conn(db_file))
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))
    monkeypatch.setattr(config, "REPORT_LABEL_MODE", "rule")
    # Disable LLM so insights / cluster deep analysis use the fallback path
    monkeypatch.setattr(config, "LLM_API_BASE", "")
    monkeypatch.setattr(config, "LLM_API_KEY", "")
    # Keep the cluster deep-analysis loop a no-op (fallback returns None anyway)
    monkeypatch.setattr(config, "REPORT_CLUSTER_ANALYSIS", False, raising=False)
    Path(config.REPORT_DIR).mkdir(parents=True, exist_ok=True)

    # init_db applies migrations 0010 + 0011 (F011 schema).
    models.init_db()

    conn = _get_test_conn(db_file)

    # 5 own + 3 competitor products. Use SCRAPED_AT inside the 2026-04-27 window
    # so query_report_data picks them up.
    products_seed = [
        # (url, site, name, sku, price, stock, review_count, rating, ownership)
        ("http://own/1", "waltons",       "Walton's #22 Meat Grinder",  "OWN-001", 729.99, "in_stock", 50, 4.8, "own"),
        ("http://own/2", "waltons",       ".75 HP Grinder",             "OWN-002", 519.99, "in_stock", 30, 4.5, "own"),
        ("http://own/3", "waltons",       "Walton's General Lug",       "OWN-003",  89.99, "in_stock", 40, 4.7, "own"),
        ("http://own/4", "waltons",       "Walton's Quick Patty Maker", "OWN-004",  69.99, "in_stock", 25, 4.3, "own"),
        ("http://own/5", "waltons",       ".5 HP Dual Grind",           "OWN-005", 399.99, "in_stock",  0, 0.0, "own"),
        ("http://comp/a", "meatyourmaker", "Competitor A 1HP Grinder",  "COMP-A",  389.99, "in_stock", 100, 4.0, "competitor"),
        ("http://comp/b", "basspro",       "Competitor B Stuffer",      "COMP-B",  249.99, "in_stock",  80, 3.8, "competitor"),
        ("http://comp/c", "basspro",       "Competitor C Mixer",        "COMP-C",  149.99, "in_stock",  60, 4.2, "competitor"),
    ]
    for (url, site, name, sku, price, stock, rcount, rating, ownership) in products_seed:
        conn.execute(
            "INSERT INTO products (url, site, name, sku, price, stock_status, "
            "review_count, rating, ownership, scraped_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (url, site, name, sku, price, stock, rcount, rating, ownership,
             "2026-04-27 09:00:00"),
        )

    # Map sku -> id for review insertion
    sku_to_id = {}
    for row in conn.execute("SELECT id, sku FROM products").fetchall():
        sku_to_id[row["sku"]] = row["id"]

    # Reviews: ~22 own (mostly negative for OWN-001, mixed elsewhere), 8 competitor (positive)
    # Each review has a corresponding analysis row. Mix of impact_category / failure_mode.
    reviews_seed = [
        # (sku, author, headline, body, body_hash, rating, date_published,
        #   sentiment, sentiment_score, labels_json, features_json, insight_cn,
        #   impact_category, failure_mode)
        ("OWN-001", "John",  "Handle broke",        "Handle came off after 3 months",   "h001", 1.0, "2026-04-15", "negative", 0.10,
         '[{"code":"quality_stability","polarity":"negative","severity":"high","confidence":0.95}]',
         '["手柄松动","手柄脱落"]', "手柄使用3个月后断裂", "durability", "structural_break"),
        ("OWN-001", "Mike",  "Terrible quality",    "The handle is loose from day one", "h002", 1.0, "2026-04-16", "negative", 0.08,
         '[{"code":"quality_stability","polarity":"negative","severity":"high","confidence":0.92}]',
         '["手柄松动"]', "手柄第一天就松了", "durability", "loose_assembly"),
        ("OWN-001", "Jane",  "Leaking everywhere",  "Seal leaks during stuffing",       "h003", 2.0, "2026-04-10", "negative", 0.15,
         '[{"code":"quality_stability","polarity":"negative","severity":"medium","confidence":0.88}]',
         '["密封漏油"]', "灌肉时密封漏油", "leakage", "seal_degradation"),
        ("OWN-001", "Pete",  "Motor died",          "Motor stopped after first use",    "h004", 1.0, "2026-04-08", "negative", 0.09,
         '[{"code":"motor_failure","polarity":"negative","severity":"high","confidence":0.91}]',
         '["电机故障"]', "电机首次使用就坏", "motor", "motor_burnout"),
        ("OWN-001", "Lana",  "Great grinder!",      "Works perfectly out of the box",   "h005", 5.0, "2026-04-05", "positive", 0.92,
         '[{"code":"strong_performance","polarity":"positive","severity":"low","confidence":0.90}]',
         '["性能强劲"]', "开箱性能强劲", None, None),
        ("OWN-002", "Bob",   "Seal problem",        "Bottom seal leaks oil",            "h006", 2.0, "2026-04-12", "negative", 0.20,
         '[{"code":"quality_stability","polarity":"negative","severity":"medium","confidence":0.85}]',
         '["密封漏油"]', "底部密封漏油", "leakage", "seal_degradation"),
        ("OWN-002", "Cara",  "Solid build",         "Heavy and well-built",             "h007", 5.0, "2026-04-09", "positive", 0.93,
         '[{"code":"solid_build","polarity":"positive","severity":"low","confidence":0.92}]',
         '["做工扎实"]', "做工很扎实", None, None),
        ("OWN-002", "Dan",   "Easy to use",         "Setup was very simple",            "h008", 4.0, "2026-04-07", "positive", 0.85,
         '[{"code":"easy_to_use","polarity":"positive","severity":"low","confidence":0.88}]',
         '["操作简单"]', "操作简单", None, None),
        ("OWN-003", "Eva",   "Lug works fine",      "Standard meat lug, no issues",     "h009", 4.0, "2026-04-11", "positive", 0.80,
         '[{"code":"quality_stability","polarity":"positive","severity":"low","confidence":0.80}]',
         '["质量稳定"]', "质量稳定", None, None),
        ("OWN-003", "Fay",   "Cracked on arrival",  "Box arrived with crack on lug",    "h010", 2.0, "2026-04-13", "negative", 0.30,
         '[{"code":"shipping_damage","polarity":"negative","severity":"low","confidence":0.78}]',
         '["运输破损"]', "运输导致破损", "packaging", "shipping_damage"),
        ("OWN-004", "Greg",  "Patties too thin",    "Hard to make thick patties",       "h011", 3.0, "2026-04-14", "negative", 0.40,
         '[{"code":"design_limitation","polarity":"negative","severity":"low","confidence":0.82}]',
         '["设计局限"]', "厚饼难做", "design", "design_limitation"),
        ("OWN-004", "Hank",  "Works great",         "Perfect for burger night",         "h012", 5.0, "2026-04-06", "positive", 0.94,
         '[{"code":"easy_to_use","polarity":"positive","severity":"low","confidence":0.90}]',
         '["易于使用"]', "汉堡之夜利器", None, None),
        ("COMP-A", "Tom",    "Built like a tank",   "Incredibly solid build quality",   "h013", 5.0, "2026-04-10", "positive", 0.95,
         '[{"code":"solid_build","polarity":"positive","severity":"low","confidence":0.95}]',
         '["做工坦克级"]', "做工非常扎实", None, None),
        ("COMP-A", "Sam",    "Powerful motor",      "Motor handles everything",         "h014", 5.0, "2026-04-09", "positive", 0.93,
         '[{"code":"strong_performance","polarity":"positive","severity":"low","confidence":0.92}]',
         '["电机强劲"]', "电机处理能力强", None, None),
        ("COMP-A", "Una",    "Worth the price",     "Premium feel, premium price",      "h015", 5.0, "2026-04-08", "positive", 0.91,
         '[{"code":"value_for_money","polarity":"positive","severity":"low","confidence":0.85}]',
         '["物有所值"]', "物有所值", None, None),
        ("COMP-B", "Lisa",   "Great stuffer",       "Stuffs sausages quickly",          "h016", 5.0, "2026-04-07", "positive", 0.92,
         '[{"code":"strong_performance","polarity":"positive","severity":"low","confidence":0.88}]',
         '["灌制速度快"]', "灌制速度快", None, None),
        ("COMP-B", "Marc",   "Good value",          "Best stuffer for the price",       "h017", 4.0, "2026-04-06", "positive", 0.86,
         '[{"code":"value_for_money","polarity":"positive","severity":"low","confidence":0.82}]',
         '["性价比高"]', "性价比高", None, None),
        ("COMP-C", "Nina",   "Smooth mixer",        "Mixes evenly without splatter",    "h018", 5.0, "2026-04-05", "positive", 0.93,
         '[{"code":"strong_performance","polarity":"positive","severity":"low","confidence":0.90}]',
         '["搅拌均匀"]', "搅拌均匀", None, None),
        ("COMP-C", "Owen",   "Easy to clean",       "Comes apart for cleaning",         "h019", 4.0, "2026-04-04", "positive", 0.84,
         '[{"code":"easy_to_use","polarity":"positive","severity":"low","confidence":0.85}]',
         '["易于清洁"]', "易于清洁", None, None),
        ("COMP-C", "Pia",    "Quiet operation",     "Doesn't wake the neighbors",       "h020", 5.0, "2026-04-03", "positive", 0.90,
         '[{"code":"quiet_operation","polarity":"positive","severity":"low","confidence":0.88}]',
         '["静音"]', "运行静音", None, None),
    ]
    for r in reviews_seed:
        (sku, author, headline, body, body_hash, rating, date_pub,
         sentiment, score, labels, features, insight_cn,
         impact_category, failure_mode) = r
        cursor = conn.execute(
            "INSERT INTO reviews (product_id, author, headline, body, body_hash, "
            "rating, date_published, images, scraped_at, "
            "headline_cn, body_cn, translate_status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (sku_to_id[sku], author, headline, body, body_hash, rating, date_pub,
             json.dumps([]), "2026-04-27 09:30:00",
             headline, body, "done"),
        )
        review_id = cursor.lastrowid
        conn.execute(
            "INSERT INTO review_analysis "
            "(review_id, sentiment, sentiment_score, labels, features, insight_cn, insight_en, "
            " impact_category, failure_mode, prompt_version, llm_model, analyzed_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'v3', 'gpt-4o-mini', '2026-04-27 10:00:00')",
            (review_id, sentiment, score, labels, features, insight_cn, insight_cn,
             impact_category, failure_mode),
        )
    conn.commit()
    conn.close()

    run = models.create_workflow_run(
        {
            "workflow_type": "daily",
            "status": "reporting",
            "logical_date": "2026-04-27",
            "trigger_key": "daily:2026-04-27:e2e",
            "data_since": "2026-04-27T00:00:00+08:00",
            "data_until": "2026-04-28T00:00:00+08:00",
            "requested_by": "e2e-test",
            "service_version": "test",
        }
    )
    return {"db_file": db_file, "run_id": run["id"], "tmp_path": tmp_path}


# ── Helpers ─────────────────────────────────────────────────────────────────


def _trigger_full_report(e2e_db: dict) -> int:
    """Run the freeze→generate pipeline and return the run_id."""
    from qbu_crawler.server.report_snapshot import (
        freeze_report_snapshot,
        generate_full_report_from_snapshot,
        load_report_snapshot,
    )

    run_id = e2e_db["run_id"]
    frozen = freeze_report_snapshot(run_id, now="2026-04-27T12:00:00+08:00")
    snapshot = load_report_snapshot(frozen["snapshot_path"])
    # send_email=False to skip SMTP; the email_body artifact is conditional on
    # send_email=True so we exercise snapshot/analytics/xlsx/html_attachment only.
    generate_full_report_from_snapshot(snapshot, send_email=False)
    return run_id


def _get_artifact_path(e2e_db: dict, run_id: int, artifact_type: str) -> str:
    conn = _get_test_conn(e2e_db["db_file"])
    try:
        arts = list_artifacts(conn, run_id=run_id)
    finally:
        conn.close()
    matches = [a for a in arts if a["artifact_type"] == artifact_type]
    assert matches, f"no artifact of type {artifact_type!r} recorded for run {run_id}"
    # If multiple, take latest by id
    return matches[-1]["path"]


# ── Tests ───────────────────────────────────────────────────────────────────


def test_e2e_full_report_generates_expected_artifacts(e2e_db):
    """All artifact types written by the full-report path are recorded.

    With ``send_email=False`` we expect snapshot + analytics + xlsx +
    html_attachment.  email_body is gated on send_email=True.
    """
    run_id = _trigger_full_report(e2e_db)

    conn = _get_test_conn(e2e_db["db_file"])
    try:
        artifacts = list_artifacts(conn, run_id=run_id)
    finally:
        conn.close()

    types = {a["artifact_type"] for a in artifacts}
    assert "snapshot" in types
    assert "analytics" in types
    assert "html_attachment" in types
    assert "xlsx" in types
    # email_body is only recorded when send_email=True
    assert "email_body" not in types


def test_e2e_excel_has_4_sheets_no_legacy(e2e_db):
    """AC-13 — Excel only contains the 4 new sheets; legacy sheets removed."""
    run_id = _trigger_full_report(e2e_db)
    excel_path = _get_artifact_path(e2e_db, run_id, "xlsx")

    wb = openpyxl.load_workbook(excel_path)
    assert sorted(wb.sheetnames) == sorted(
        ["核心数据", "行动建议", "评论原文", "竞品启示"]
    )
    # Legacy sheets must not appear
    assert "今日变化" not in wb.sheetnames
    assert "问题标签" not in wb.sheetnames
    assert "趋势数据" not in wb.sheetnames


def test_e2e_core_data_sheet_unified_denominator(e2e_db):
    """AC-1 — 核心数据 has both denominator columns + 状态灯 (H10)."""
    run_id = _trigger_full_report(e2e_db)
    excel_path = _get_artifact_path(e2e_db, run_id, "xlsx")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["核心数据"]
    headers = [c.value for c in ws[1]]
    assert "差评率(采集分母)" in headers
    assert "差评率(站点分母)" in headers
    assert "状态灯" in headers


def test_e2e_attachment_html_under_1mb(e2e_db):
    """AC-27 — Attachment HTML ≤ 1MB."""
    run_id = _trigger_full_report(e2e_db)
    html_path = _get_artifact_path(e2e_db, run_id, "html_attachment")
    size = os.path.getsize(html_path)
    assert size <= 1 * 1024 * 1024, f"attachment HTML is {size} bytes (limit 1MB)"


def test_e2e_excel_under_5mb(e2e_db):
    """AC-27 — Excel ≤ 5MB."""
    run_id = _trigger_full_report(e2e_db)
    excel_path = _get_artifact_path(e2e_db, run_id, "xlsx")
    size = os.path.getsize(excel_path)
    assert size <= 5 * 1024 * 1024, f"Excel is {size} bytes (limit 5MB)"


def test_e2e_report_generation_under_30s(e2e_db):
    """AC-19 — Full report generation completes in < 30s on synthetic fixture."""
    start = time.time()
    _trigger_full_report(e2e_db)
    elapsed = time.time() - start
    assert elapsed < 30.0, f"Report took {elapsed:.1f}s (limit 30s)"


def test_e2e_attachment_html_no_engineering_signals(e2e_db):
    """F011 §4.1.3 — Attachment HTML must not leak raw engineering tokens.

    These specific tokens are internal field names that must never appear
    verbatim in the user-facing HTML (separate from translations of the
    underlying concepts).
    """
    run_id = _trigger_full_report(e2e_db)
    html_path = _get_artifact_path(e2e_db, run_id, "html_attachment")
    html = Path(html_path).read_text(encoding="utf-8")

    assert "estimated_dates" not in html
    assert "backfill_dominant" not in html
