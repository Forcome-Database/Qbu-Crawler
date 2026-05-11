import copy
import json
from pathlib import Path

import openpyxl

from qbu_crawler import config
from qbu_crawler.server.report import generate_excel, render_email_full
from qbu_crawler.server.report_contract import build_report_user_contract
from qbu_crawler.server.report_html import _render_v3_html_string


FIXTURE_DIR = Path(__file__).parents[1] / "fixtures" / "report_replay"


def _load_fixture():
    snapshot = json.loads((FIXTURE_DIR / "test7_minimal_snapshot.json").read_text(encoding="utf-8"))
    analytics = json.loads((FIXTURE_DIR / "test7_minimal_analytics.json").read_text(encoding="utf-8"))
    return snapshot, analytics


def _with_contract(snapshot, analytics):
    enriched = copy.deepcopy(analytics)
    enriched["report_user_contract"] = build_report_user_contract(
        snapshot=snapshot,
        analytics=enriched,
    )
    return enriched


def test_test7_replay_contract_preserves_user_semantics():
    snapshot, analytics = _load_fixture()

    contract = build_report_user_contract(snapshot=snapshot, analytics=analytics)

    actions = [item["full_action"] for item in contract["action_priorities"]]
    assert len(set(actions)) == len(actions)
    assert any(card["image_evidence"] for card in contract["issue_diagnostics"])
    assert contract["competitor_insights"]["avoid_competitor_failures"]
    assert contract["delivery"]["workflow_notification_delivered"] is False


def test_test7_replay_html_excel_email_consume_contract(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    snapshot, analytics = _load_fixture()
    analytics = _with_contract(snapshot, analytics)

    html = _render_v3_html_string(snapshot, analytics)
    assert "issue-image-evidence" in html
    assert "https://example.test/review-101.jpg" in html
    assert "复核成型尺寸并验证肉饼成型路径" in html
    assert "data-product=\"Walton&#39;s Quick Patty Maker\"" in html or "data-product=\"Walton's Quick Patty Maker\"" in html

    excel_path = generate_excel(
        snapshot["products"],
        snapshot["reviews"],
        analytics=analytics,
        output_path=str(tmp_path / "test7-replay.xlsx"),
    )
    wb = openpyxl.load_workbook(excel_path)
    actions_ws = wb["行动建议"]
    action_headers = [cell.value for cell in actions_ws[1]]
    full_action_col = action_headers.index("改良方向") + 1
    full_actions = [
        actions_ws.cell(row=row, column=full_action_col).value
        for row in range(2, actions_ws.max_row + 1)
        if actions_ws.cell(row=row, column=full_action_col).value
    ]
    assert len(set(full_actions)) == len(full_actions)

    competitor_ws = wb["竞品启示"]
    competitor_headers = [cell.value for cell in competitor_ws[1]]
    competitor_values = [cell.value for row in competitor_ws.iter_rows() for cell in row]
    assert "对自有产品启发" in competitor_headers
    assert "样本数" in competitor_headers
    assert "涉及产品数" in competitor_headers
    assert "竞品包装容易破损" in competitor_values

    email_html = render_email_full(snapshot, analytics)
    # 新版邮件 label / count 分列 — 分别断言而非连续字符串
    assert "需关注产品" in email_html
    assert ">2<" in email_html
    assert "复核结构尺寸" in email_html
