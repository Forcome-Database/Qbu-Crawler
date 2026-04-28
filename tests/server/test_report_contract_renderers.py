import openpyxl

from qbu_crawler import config
from qbu_crawler.server.report import generate_excel, render_email_full
from qbu_crawler.server.report_common import normalize_deep_report_analytics
from qbu_crawler.server.report_html import _render_v3_html_string
from qbu_crawler.server.report_snapshot import _merge_post_normalize_mutations


def _contract():
    return {
        "schema_version": "report_user_contract.v1",
        "contract_context": {"snapshot_source": "provided"},
        "kpis": {
            "health_index": 91.2,
            "own_positive_review_rows": 8,
            "own_review_rows": 10,
            "own_negative_review_rate": 0.1,
            "own_negative_review_rate_display": "10.0%",
            "own_product_count": 1,
            "competitor_product_count": 0,
            "ingested_review_rows": 10,
            "attention_product_count": 1,
        },
        "action_priorities": [{
            "label_code": "structure_design",
            "label_display": "结构设计",
            "short_title": "复核结构尺寸",
            "full_action": "复核结构尺寸并验证肉饼成型路径",
            "affected_products": ["Walton's Quick Patty Maker"],
            "affected_products_count": 1,
            "top_complaint": "肉饼太大",
            "evidence_count": 1,
            "evidence_review_ids": [101],
            "source": "llm_rewrite",
        }],
        "issue_diagnostics": [{
            "label_code": "structure_design",
            "label_display": "结构设计",
            "feature_display": "结构设计",
            "severity": "high",
            "default_expanded": True,
            "affected_products": ["Walton's Quick Patty Maker"],
            "evidence_count": 1,
            "evidence_review_ids": [101],
            "text_evidence": [{"review_id": 101, "display_body": "肉饼太大"}],
            "example_reviews": [{"id": 101, "summary_text": "肉饼太大"}],
            "image_evidence": [{"review_id": 101, "url": "https://example.test/a.jpg"}],
            "ai_recommendation": "复核结构尺寸并验证肉饼成型路径",
            "recommended_action": "复核结构尺寸并验证肉饼成型路径",
            "failure_modes": [{"name": "尺寸不匹配"}],
            "root_causes": [{"name": "模具约束不足"}],
            "user_workarounds": ["手工修整"],
        }],
        "heatmap": {
            "x_labels": ["结构设计"],
            "x_label_codes": ["structure_design"],
            "y_labels": ["Walton's Quick Patty"],
            "y_items": [{"product_name": "Walton's Quick Patty Maker", "display_label": "Walton's Quick Patty"}],
            "z": [[{
                "score": 0.5,
                "score_display": "50%",
                "color_class": "heatmap-mid",
                "tooltip": "体验健康度 50%",
                "top_review_excerpt": "肉饼太大",
            }]],
        },
    }


def test_excel_actions_can_render_from_contract_only(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    analytics = {"report_user_contract": _contract(), "kpis": {}, "self": {}, "competitor": {}}

    path = generate_excel(
        [{"name": "Walton's Quick Patty Maker", "sku": "WQPM", "ownership": "own"}],
        [],
        analytics=analytics,
        output_path=str(tmp_path / "contract.xlsx"),
    )

    wb = openpyxl.load_workbook(path)
    ws = wb["现在该做什么"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "复核结构尺寸" in values
    assert "复核结构尺寸并验证肉饼成型路径" in values


def test_html_issue_cards_can_render_from_contract_only():
    snapshot = {
        "logical_date": "2026-04-28",
        "products": [{"name": "Walton's Quick Patty Maker"}],
        "reviews": [],
    }
    analytics = {"report_semantics": "bootstrap", "report_user_contract": _contract()}

    html = _render_v3_html_string(snapshot, analytics)

    assert "issue-image-evidence" in html
    assert "复核结构尺寸并验证肉饼成型路径" in html
    assert "尺寸不匹配" in html


def test_email_top_actions_can_render_from_contract_only():
    snapshot = {"logical_date": "2026-04-28"}
    analytics = {"report_user_contract": _contract()}

    html = render_email_full(snapshot, analytics)

    assert "复核结构尺寸" in html
    assert "需关注产品 1 个" in html


def test_html_refreshes_missing_snapshot_contract(monkeypatch):
    calls = []

    def fake_build_contract(*, snapshot, analytics, llm_copy=None):
        calls.append(snapshot)
        contract = _contract()
        contract["contract_context"] = {"snapshot_source": "provided"}
        return contract

    monkeypatch.setattr(
        "qbu_crawler.server.report_html.build_report_user_contract",
        fake_build_contract,
    )
    snapshot = {
        "logical_date": "2026-04-28",
        "products": [{"name": "Walton's Quick Patty Maker"}],
        "reviews": [],
    }
    analytics = {
        "report_semantics": "bootstrap",
        "report_user_contract": {
            "schema_version": "report_user_contract.v1",
            "contract_context": {"snapshot_source": "missing"},
            "issue_diagnostics": [{"label_display": "stale temporary card"}],
        },
    }

    html = _render_v3_html_string(snapshot, analytics)

    assert calls and calls[0] == snapshot
    assert "复核结构尺寸并验证肉饼成型路径" in html
    assert "stale temporary card" not in html


def test_excel_competitor_insights_can_render_from_contract_only(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    contract = _contract()
    contract["competitor_insights"] = {
        "learn_from_competitors": [{
            "summary_cn": "竞品安装步骤清晰",
            "self_product_implication": "自有产品需要复核说明书步骤表达",
            "suggested_validation": "抽查安装说明和客服问答",
            "evidence_review_ids": [201],
            "sample_size": 8,
            "product_count": 2,
        }],
        "avoid_competitor_failures": [{
            "summary_cn": "竞品包装容易破损",
            "self_product_implication": "自有产品需压测包装边角保护",
            "suggested_validation": "做一次运输跌落复测",
            "evidence_review_ids": [301],
            "sample_size": 3,
            "product_count": 1,
        }],
        "validation_hypotheses": [{
            "summary_cn": "验证包装改良能否降低破损投诉",
            "self_product_implication": "优先覆盖高退货风险 SKU",
            "suggested_validation": "小批量 A/B 包装验证",
            "evidence_review_ids": [301],
            "sample_size": 3,
            "product_count": 1,
        }],
    }
    analytics = {"report_user_contract": contract, "kpis": {}, "self": {}, "competitor": {}}

    path = generate_excel(
        [{"name": "Walton's Quick Patty Maker", "sku": "WQPM", "ownership": "own"}],
        [],
        analytics=analytics,
        output_path=str(tmp_path / "contract-competitor.xlsx"),
    )

    wb = openpyxl.load_workbook(path)
    ws = wb["竞品启示"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    headers = [cell.value for cell in ws[1]]
    assert "对自有产品启发" in headers
    assert "验证动作" in headers
    assert "样本数" in headers
    assert "涉及产品数" in headers
    assert "竞品安装步骤清晰" in values
    assert "自有产品需压测包装边角保护" in values


def test_html_bootstrap_changes_tab_reads_contract_digest():
    snapshot = {
        "logical_date": "2026-04-28",
        "products": [{"name": "A"}, {"name": "B"}],
        "reviews": [{"id": 1}, {"id": 2}, {"id": 3}],
    }
    contract = _contract()
    contract["bootstrap_digest"] = {
        "baseline_summary": {
            "headline": "监控起点已建立",
            "product_count": 2,
            "review_count": 3,
            "coverage_rate": 0.75,
            "translation_completion_rate": 1.0,
        },
        "data_quality": {
            "historical_backfill_ratio": 0.8,
            "estimated_date_ratio": 0.2,
            "low_coverage_products": ["B"],
        },
        "immediate_attention": ["需关注产品 2 个"],
    }
    analytics = {"report_semantics": "bootstrap", "report_user_contract": contract}

    html = _render_v3_html_string(snapshot, analytics)

    assert "监控起点已建立" in html
    assert "当前截面：2 款产品 / 3 条评论" in html
    assert "低覆盖产品：B" in html
    assert "需关注产品 2 个" in html
    assert "较昨日" not in html


def test_post_normalize_mutations_refresh_contract_with_real_snapshot():
    normalized = normalize_deep_report_analytics({
        "report_semantics": "bootstrap",
        "kpis": {"health_index": 96.2},
    })
    raw = {
        "report_semantics": "bootstrap",
        "report_copy": {
            "hero_headline": "Health index 96.2 is stable",
            "executive_summary": "Summary",
            "executive_bullets": [],
            "improvement_priorities": [],
        },
    }
    snapshot = {
        "logical_date": "2026-04-28",
        "products": [{"name": "Walton's Quick Patty Maker"}],
        "reviews": [{"id": 101}],
    }

    _merge_post_normalize_mutations(normalized, raw, snapshot=snapshot)

    context = normalized["report_user_contract"]["contract_context"]
    assert context["snapshot_source"] == "provided"
    assert context["product_count"] == 1
    assert context["review_count"] == 1
