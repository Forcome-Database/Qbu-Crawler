from pathlib import Path

import openpyxl

from qbu_crawler.server.report import generate_excel, render_email_full
from qbu_crawler.server.report_common import normalize_deep_report_analytics
from qbu_crawler.server.report_html import _render_v3_html_string


def test_strict_mode_adapts_legacy_analytics_once():
    analytics = {
        "report_semantics": "bootstrap",
        "self": {
            "top_negative_clusters": [{
                "label_code": "structure_design",
                "label_display": "结构设计",
                "review_count": 3,
                "example_reviews": [{"id": 1, "body_cn": "尺寸不合适"}],
            }]
        },
        "report_copy": {
            "improvement_priorities": [{
                "label_code": "structure_design",
                "full_action": "复核结构尺寸",
                "evidence_review_ids": [1],
                "affected_products": ["A"],
            }]
        },
    }

    normalized = normalize_deep_report_analytics(analytics)

    contract = normalized["report_user_contract"]
    assert contract["contract_source"] == "legacy_adapter"
    assert contract["action_priorities"][0]["full_action"] == "复核结构尺寸"
    assert contract["issue_diagnostics"][0]["label_code"] == "structure_design"


def test_strict_mode_hides_legacy_action_without_evidence():
    analytics = {
        "report_semantics": "bootstrap",
        "report_copy": {
            "improvement_priorities": [{
                "label_code": "generic",
                "full_action": "补充检查",
                "affected_products": ["A"],
            }]
        },
    }

    normalized = normalize_deep_report_analytics(analytics)

    assert normalized["report_user_contract"]["action_priorities"] == []


def test_non_strict_mode_keeps_legacy_action_without_evidence(monkeypatch):
    from qbu_crawler import config

    monkeypatch.setattr(config, "REPORT_CONTRACT_STRICT_MODE", False)
    analytics = {
        "report_semantics": "bootstrap",
        "report_copy": {
            "improvement_priorities": [{
                "label_code": "generic",
                "full_action": "补充检查",
                "affected_products": ["A"],
            }]
        },
    }

    normalized = normalize_deep_report_analytics(analytics)

    assert normalized["report_user_contract"]["action_priorities"][0]["full_action"] == "补充检查"


def test_business_html_does_not_render_ops_diagnostics():
    analytics = {
        "report_semantics": "bootstrap",
        "report_user_contract": {
            "bootstrap_digest": {
                "baseline_summary": {"headline": "监控起点已建立", "product_count": 1, "review_count": 2},
                "immediate_attention": [],
            },
            "delivery": {
                "deadletter_count": 3,
                "workflow_notification_delivered": False,
            },
        },
        "data_quality": {"low_coverage_products": ["SKU-X"], "estimated_date_ratio": 0.5},
    }

    html = _render_v3_html_string({"logical_date": "2026-04-28", "products": [], "reviews": []}, analytics)

    assert "deadletter" not in html
    assert "SKU-X" not in html
    assert "estimated_date_ratio" not in html


def test_v3_template_does_not_render_ops_diagnostics():
    text = Path("qbu_crawler/server/report_templates/daily_report_v3.html.j2").read_text(encoding="utf-8")
    assert "low_coverage_products" not in text
    assert "deadletter_count" not in text
    assert "estimated_date_ratio" not in text


def test_strict_excel_does_not_render_legacy_action_without_evidence(tmp_path, monkeypatch):
    from qbu_crawler import config

    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    analytics = {
        "report_semantics": "bootstrap",
        "report_copy": {
            "improvement_priorities": [{
                "label_code": "generic",
                "short_title": "补充检查",
                "full_action": "补充检查",
                "affected_products": ["A"],
            }]
        },
    }

    path = generate_excel(
        [{"name": "A", "sku": "A", "ownership": "own"}],
        [],
        analytics=analytics,
        output_path=str(tmp_path / "strict.xlsx"),
    )

    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[1]]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "补充检查" not in values


def test_strict_email_does_not_render_legacy_action_without_evidence():
    analytics = {
        "report_semantics": "bootstrap",
        "kpis": {"own_product_count": 1, "competitor_product_count": 0, "ingested_review_rows": 0},
        "report_copy": {
            "improvement_priorities": [{
                "label_code": "generic",
                "label_display": "通用",
                "short_title": "补充检查",
                "full_action": "补充检查",
                "affected_products": ["A"],
                "affected_products_count": 1,
            }]
        },
    }

    html = render_email_full({"logical_date": "2026-04-28"}, analytics)

    assert "补充检查" not in html
