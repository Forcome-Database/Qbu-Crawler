"""Tests for server/report.py"""

import json
import os
import smtplib
import tempfile
from datetime import datetime, timezone, timedelta
from email import message_from_string
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from qbu_crawler import models, config


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def patch_db(tmp_path, monkeypatch):
    """Create a temp SQLite DB, init schema, insert one product + one review."""
    db_file = str(tmp_path / "test_products.db")
    monkeypatch.setattr(config, "DB_PATH", db_file)
    monkeypatch.setattr(models, "get_conn", lambda: _get_test_conn(db_file))

    models.init_db()

    conn = _get_test_conn(db_file)
    # Insert a product (scraped_at = now)
    conn.execute(
        """
        INSERT INTO products (url, site, name, sku, price, stock_status,
                              review_count, rating, ownership, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """,
        (
            "https://example.com/product/1",
            "basspro",
            "Test Product",
            "SKU-001",
            49.99,
            "in_stock",
            5,
            4.5,
            "own",
        ),
    )
    conn.commit()

    product_id = conn.execute("SELECT id FROM products WHERE sku = 'SKU-001'").fetchone()["id"]

    images_json = json.dumps(["https://img.example.com/1.jpg"])
    conn.execute(
        """
        INSERT INTO reviews (product_id, author, headline, body, body_hash,
                             rating, date_published, images, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """,
        (
            product_id,
            "John Doe",
            "Great product",
            "Really loved it, would buy again.",
            "abc123",
            5.0,
            "2026-03-01",
            images_json,
        ),
    )
    conn.commit()
    conn.close()

    yield db_file


def _get_test_conn(db_file: str):
    """Return a sqlite3 connection with Row factory pointing at db_file."""
    import sqlite3
    conn = sqlite3.connect(db_file)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------------------------
# query_report_data
# ---------------------------------------------------------------------------

def test_query_report_data(patch_db, monkeypatch):
    monkeypatch.setattr(config, "DB_PATH", patch_db)

    from qbu_crawler.server.report import query_report_data

    # Query from 1 hour ago — should capture the test data
    since = datetime.now(timezone.utc) - timedelta(hours=1)
    products, reviews = query_report_data(since)

    assert len(products) == 1
    assert products[0]["name"] == "Test Product"
    assert products[0]["sku"] == "SKU-001"
    assert products[0]["site"] == "basspro"
    assert products[0]["ownership"] == "own"

    assert len(reviews) == 1
    assert reviews[0]["author"] == "John Doe"
    assert reviews[0]["headline"] == "Great product"
    assert reviews[0]["product_name"] == "Test Product"
    # images should be parsed from JSON
    assert isinstance(reviews[0]["images"], list)
    assert reviews[0]["images"][0] == "https://img.example.com/1.jpg"


def test_query_report_data_future_cutoff(patch_db, monkeypatch):
    monkeypatch.setattr(config, "DB_PATH", patch_db)

    from qbu_crawler.server.report import query_report_data

    # Query from 1 hour in the future — should return nothing
    since = datetime.now(timezone.utc) + timedelta(hours=1)
    products, reviews = query_report_data(since)

    assert products == []
    assert reviews == []


def test_query_report_data_respects_until(tmp_path, monkeypatch):
    db_file = str(tmp_path / "bounded-window.db")
    monkeypatch.setattr(config, "DB_PATH", db_file)
    monkeypatch.setattr(models, "get_conn", lambda: _get_test_conn(db_file))
    models.init_db()

    conn = _get_test_conn(db_file)
    conn.execute(
        """
        INSERT INTO products (url, site, name, sku, price, stock_status,
                              review_count, rating, ownership, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "https://example.com/product/in-window",
            "basspro",
            "Window Product",
            "SKU-001",
            49.99,
            "in_stock",
            1,
            4.5,
            "own",
            "2026-03-27 08:30:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO products (url, site, name, sku, price, stock_status,
                              review_count, rating, ownership, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "https://example.com/product/out-window",
            "basspro",
            "Future Product",
            "SKU-002",
            59.99,
            "in_stock",
            1,
            4.0,
            "competitor",
            "2026-03-27 10:30:00",
        ),
    )
    product_id = conn.execute("SELECT id FROM products WHERE sku = 'SKU-001'").fetchone()["id"]
    future_product_id = conn.execute("SELECT id FROM products WHERE sku = 'SKU-002'").fetchone()["id"]
    conn.execute(
        """
        INSERT INTO reviews (product_id, author, headline, body, body_hash,
                             rating, date_published, images, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            product_id,
            "Alice",
            "Included",
            "Inside window",
            "hash-included",
            5.0,
            "2026-03-27",
            "[]",
            "2026-03-27 08:35:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO reviews (product_id, author, headline, body, body_hash,
                             rating, date_published, images, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            future_product_id,
            "Bob",
            "Excluded",
            "Outside window",
            "hash-excluded",
            5.0,
            "2026-03-27",
            "[]",
            "2026-03-27 10:35:00",
        ),
    )
    conn.commit()
    conn.close()

    from qbu_crawler.server.report import query_report_data

    products, reviews = query_report_data(
        "2026-03-27T00:00:00+00:00",
        until="2026-03-27T02:00:00+00:00",
    )

    assert [item["sku"] for item in products] == ["SKU-001"]
    assert [item["author"] for item in reviews] == ["Alice"]


def test_query_report_data_converts_aware_window_to_shanghai_time(tmp_path, monkeypatch):
    db_file = str(tmp_path / "aware-window.db")
    monkeypatch.setattr(config, "DB_PATH", db_file)
    monkeypatch.setattr(models, "get_conn", lambda: _get_test_conn(db_file))
    models.init_db()

    conn = _get_test_conn(db_file)
    conn.execute(
        """
        INSERT INTO products (url, site, name, sku, price, stock_status,
                              review_count, rating, ownership, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "https://example.com/product/early",
            "basspro",
            "Early Product",
            "SKU-EARLY",
            39.99,
            "in_stock",
            1,
            4.0,
            "own",
            "2026-03-27 07:30:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO products (url, site, name, sku, price, stock_status,
                              review_count, rating, ownership, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "https://example.com/product/late",
            "basspro",
            "Late Product",
            "SKU-LATE",
            49.99,
            "in_stock",
            1,
            4.0,
            "own",
            "2026-03-27 08:30:00",
        ),
    )
    early_product_id = conn.execute("SELECT id FROM products WHERE sku = 'SKU-EARLY'").fetchone()["id"]
    late_product_id = conn.execute("SELECT id FROM products WHERE sku = 'SKU-LATE'").fetchone()["id"]
    conn.execute(
        """
        INSERT INTO reviews (product_id, author, headline, body, body_hash,
                             rating, date_published, images, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            early_product_id,
            "Alice",
            "Too early",
            "Should be excluded",
            "hash-early",
            5.0,
            "2026-03-27",
            "[]",
            "2026-03-27 07:35:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO reviews (product_id, author, headline, body, body_hash,
                             rating, date_published, images, scraped_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            late_product_id,
            "Bob",
            "In window",
            "Should be included",
            "hash-late",
            5.0,
            "2026-03-27",
            "[]",
            "2026-03-27 08:35:00",
        ),
    )
    conn.commit()
    conn.close()

    from qbu_crawler.server.report import query_report_data

    products, reviews = query_report_data(
        "2026-03-27T00:00:00+00:00",
        until="2026-03-28T00:00:00+00:00",
    )

    assert [item["sku"] for item in products] == ["SKU-LATE"]
    assert [item["author"] for item in reviews] == ["Bob"]


# ---------------------------------------------------------------------------
# generate_excel
# ---------------------------------------------------------------------------

def test_generate_excel(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))

    from qbu_crawler.server import report
    from qbu_crawler.server.report import generate_excel
    from openpyxl import load_workbook

    monkeypatch.setattr(report, "_download_and_resize", lambda url: None)

    products = [
        {
            "url": "https://example.com/p/1",
            "name": "Test Product",
            "sku": "SKU-001",
            "price": 49.99,
            "stock_status": "in_stock",
            "rating": 4.5,
            "review_count": 10,
            "scraped_at": "2026-03-10 00:00:00",
            "site": "basspro",
            "ownership": "own",
        }
    ]
    reviews = [
        {
            "product_name": "Test Product",
            "author": "John Doe",
            "headline": "Great",
            "body": "Loved it",
            "headline_cn": "很棒",
            "body_cn": "喜欢",
            "rating": 5.0,
            "date_published": "2026-03-01",
            "images": ["https://img.example.com/1.jpg"],
        }
    ]

    report_date = datetime(2026, 3, 10, tzinfo=timezone.utc)
    filepath = generate_excel(products, reviews, report_date=report_date)

    assert os.path.isfile(filepath)
    assert filepath.endswith("scrape-report-2026-03-10.xlsx")

    wb = load_workbook(filepath)
    assert "产品" in wb.sheetnames
    assert "评论" in wb.sheetnames

    ws_p = wb["产品"]
    # Header row
    headers = [ws_p.cell(row=1, column=c).value for c in range(1, 11)]
    assert "产品地址" in headers
    assert "产品名称" in headers
    assert "SKU" in headers

    # Data row
    assert ws_p.cell(row=2, column=1).value == "https://example.com/p/1"
    assert ws_p.cell(row=2, column=2).value == "Test Product"

    ws_r = wb["评论"]
    r_headers = [ws_r.cell(row=1, column=c).value for c in range(1, 11)]
    assert "产品名称" in r_headers
    assert "标题（中文）" in r_headers
    assert ws_r.cell(row=2, column=9).value == "2026-03-01"
    # When image downloads are unavailable, URLs fall back to the last column.
    images_cell = ws_r.cell(row=2, column=10).value
    assert "img.example.com" in images_cell


def test_generate_excel_empty(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))

    from qbu_crawler.server.report import generate_excel
    from openpyxl import load_workbook

    report_date = datetime(2026, 3, 10, tzinfo=timezone.utc)
    filepath = generate_excel([], [], report_date=report_date)

    assert os.path.isfile(filepath)
    wb = load_workbook(filepath)
    assert "产品" in wb.sheetnames
    assert "评论" in wb.sheetnames

    ws_p = wb["产品"]
    # Only header row, no data
    assert ws_p.max_row == 1
    assert ws_p.cell(row=1, column=1).value == "产品地址"

    ws_r = wb["评论"]
    assert ws_r.max_row == 1
    assert ws_r.cell(row=1, column=1).value == "产品名称"


# ---------------------------------------------------------------------------
# send_email
# ---------------------------------------------------------------------------

def test_send_email_success(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "user@example.com")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "secret")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    mock_smtp_instance = MagicMock()

    with patch("smtplib.SMTP", return_value=mock_smtp_instance) as mock_smtp_cls:
        from qbu_crawler.server.report import send_email

        result = send_email(
            recipients=["recipient@example.com"],
            subject="Test Subject",
            body_text="Hello World",
        )

    assert result["success"] is True
    assert result["error"] is None
    assert result["recipients"] == 1
    mock_smtp_cls.assert_called_once_with("smtp.example.com", 587)
    mock_smtp_instance.starttls.assert_called_once()
    mock_smtp_instance.login.assert_called_once_with("user@example.com", "secret")
    mock_smtp_instance.sendmail.assert_called_once()
    mock_smtp_instance.quit.assert_called_once()


def test_send_email_ssl(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 465)
    monkeypatch.setattr(config, "SMTP_USER", "user@example.com")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "secret")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", True)

    mock_smtp_instance = MagicMock()

    with patch("smtplib.SMTP_SSL", return_value=mock_smtp_instance):
        from qbu_crawler.server.report import send_email

        result = send_email(
            recipients=["recipient@example.com"],
            subject="Test",
            body_text="Body",
        )

    assert result["success"] is True
    # No STARTTLS for SSL connections
    mock_smtp_instance.starttls.assert_not_called()


def test_send_email_supports_multiple_attachments(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    first = tmp_path / "a.xlsx"
    second = tmp_path / "b.pdf"
    first.write_text("a", encoding="utf-8")
    second.write_text("b", encoding="utf-8")

    mock_smtp_instance = MagicMock()

    with patch("smtplib.SMTP", return_value=mock_smtp_instance):
        from qbu_crawler.server.report import send_email

        result = send_email(
            recipients=["recipient@example.com"],
            subject="Test",
            body_text="Body",
            attachment_paths=[str(first), str(second)],
        )

    assert result["success"] is True
    raw_message = mock_smtp_instance.sendmail.call_args.args[2]
    parsed = message_from_string(raw_message)
    filenames = sorted(
        part.get_filename()
        for part in parsed.walk()
        if part.get_filename()
    )
    assert filenames == ["a.xlsx", "b.pdf"]


def test_send_email_encodes_non_ascii_attachment_filename(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    attachment = tmp_path / "日报分析.pdf"
    attachment.write_text("pdf", encoding="utf-8")

    mock_smtp_instance = MagicMock()

    with patch("smtplib.SMTP", return_value=mock_smtp_instance):
        from qbu_crawler.server.report import send_email

        result = send_email(
            recipients=["recipient@example.com"],
            subject="Test",
            body_text="Body",
            attachment_path=str(attachment),
        )

    assert result["success"] is True
    raw_message = mock_smtp_instance.sendmail.call_args.args[2]
    assert "filename*=" in raw_message


def test_send_email_attachment_path_still_works(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    attachment = tmp_path / "legacy.xlsx"
    attachment.write_text("legacy", encoding="utf-8")

    mock_smtp_instance = MagicMock()

    with patch("smtplib.SMTP", return_value=mock_smtp_instance):
        from qbu_crawler.server.report import send_email

        result = send_email(
            recipients=["recipient@example.com"],
            subject="Test",
            body_text="Body",
            attachment_path=str(attachment),
        )

    assert result["success"] is True
    raw_message = mock_smtp_instance.sendmail.call_args.args[2]
    parsed = message_from_string(raw_message)
    filenames = [part.get_filename() for part in parsed.walk() if part.get_filename()]
    assert filenames == ["legacy.xlsx"]


def test_send_email_no_config(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "")

    from qbu_crawler.server.report import send_email

    result = send_email(
        recipients=["recipient@example.com"],
        subject="Test",
        body_text="Hello",
    )

    assert result["success"] is False
    assert "SMTP_HOST" in result["error"]
    assert result["recipients"] == 0


def test_send_email_no_recipients(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")

    from qbu_crawler.server.report import send_email

    result = send_email(recipients=[], subject="Test", body_text="Hello")

    assert result["success"] is False
    assert result["recipients"] == 0


def test_send_email_smtp_error(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    with patch("smtplib.SMTP", side_effect=ConnectionRefusedError("Connection refused")):
        from qbu_crawler.server.report import send_email

        result = send_email(
            recipients=["r@example.com"],
            subject="Test",
            body_text="Hello",
        )

    assert result["success"] is False
    assert "Connection refused" in result["error"]
    assert result["recipients"] == 0


def test_send_email_retries_transient_send_failure(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    smtp_instances = []

    class FakeSMTP:
        def __init__(self, host, port):
            assert host == "smtp.example.com"
            assert port == 587
            self.quit_calls = 0
            self.sendmail_calls = 0
            smtp_instances.append(self)

        def starttls(self):
            return None

        def login(self, *_args):
            return None

        def sendmail(self, *_args):
            self.sendmail_calls += 1
            if len(smtp_instances) == 1:
                raise smtplib.SMTPServerDisconnected("temporary drop")
            return None

        def quit(self):
            self.quit_calls += 1

    monkeypatch.setattr("smtplib.SMTP", FakeSMTP)
    monkeypatch.setattr("time.sleep", lambda _seconds: None)

    from qbu_crawler.server.report import send_email

    result = send_email(
        recipients=["recipient@example.com"],
        subject="Test",
        body_text="Body",
    )

    assert result["success"] is True
    assert len(smtp_instances) == 2
    assert smtp_instances[0].quit_calls == 1
    assert smtp_instances[1].quit_calls == 1


def test_send_email_closes_connection_on_sendmail_error(monkeypatch):
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "sender@example.com")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)

    smtp_instances = []

    class FakeSMTP:
        def __init__(self, *_args):
            self.quit_calls = 0
            smtp_instances.append(self)

        def starttls(self):
            return None

        def sendmail(self, *_args):
            raise smtplib.SMTPException("send failed")

        def quit(self):
            self.quit_calls += 1

    monkeypatch.setattr("smtplib.SMTP", FakeSMTP)
    monkeypatch.setattr("time.sleep", lambda _seconds: None)

    from qbu_crawler.server.report import send_email

    result = send_email(
        recipients=["recipient@example.com"],
        subject="Test",
        body_text="Body",
    )

    assert result["success"] is False
    assert len(smtp_instances) == 3
    assert all(item.quit_calls == 1 for item in smtp_instances)


# ---------------------------------------------------------------------------
# load_email_recipients
# ---------------------------------------------------------------------------

def test_load_email_recipients(tmp_path):
    from qbu_crawler.server.report import load_email_recipients

    recipients_file = tmp_path / "recipients.txt"
    recipients_file.write_text(
        "# Comment line\n"
        "alice@example.com\n"
        "bob@example.com\n"
        "\n"
        "# Another comment\n"
        "carol@example.com\n",
        encoding="utf-8",
    )

    result = load_email_recipients(str(recipients_file))
    assert result == ["alice@example.com", "bob@example.com", "carol@example.com"]


def test_load_email_recipients_missing_file():
    from qbu_crawler.server.report import load_email_recipients

    result = load_email_recipients("/nonexistent/path/recipients.txt")
    assert result == []


def test_build_daily_deep_report_email_keeps_only_core_summary():
    from qbu_crawler.server.report import build_daily_deep_report_email

    subject, body = build_daily_deep_report_email(
        {
            "logical_date": "2026-04-03",
            "data_until": "2026-04-04T00:00:00+08:00",
        },
        {
            "mode": "incremental",
            "report_copy": {
                "executive_bullets": [
                    "自有产品 Own Grinder 差评集中在质量稳定性。",
                    "竞品 Competitor Grinder 的做工口碑持续稳定。",
                    "图片证据已覆盖本期主要风险样本。",
                ]
            },
            "kpis": {
                "product_count": 12,
                "ingested_review_rows": 186,
                "translated_count": 180,
                "untranslated_count": 6,
            },
        },
    )

    assert subject
    assert "产品评论日报" in subject
    assert "2026-04-03" in subject
    assert "今日要点：" in body
    assert "详见附件 PDF" in body
    assert "自有产品重点风险" not in body
    assert "问题簇与改良方向" not in body
    assert "竞品机会窗口" not in body


# ---------------------------------------------------------------------------
# generate_report (full pipeline)
# ---------------------------------------------------------------------------

def test_generate_report_full(patch_db, tmp_path, monkeypatch):
    """Full pipeline — translation comes from DB, not inline."""
    monkeypatch.setattr(config, "DB_PATH", patch_db)
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))

    from qbu_crawler.server.report import generate_report

    since = datetime.now(timezone.utc) - timedelta(hours=1)
    result = generate_report(since, send_email=False)

    assert result["products_count"] == 1
    assert result["reviews_count"] == 1
    assert result["translated_count"] == 0
    assert result["untranslated_count"] == 1
    assert os.path.isfile(result["excel_path"])
    assert result["email"] is None


def test_generate_report_with_pretranslated_reviews(patch_db, tmp_path, monkeypatch):
    """Pipeline reads pre-translated data from DB."""
    monkeypatch.setattr(config, "DB_PATH", patch_db)
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))

    import sqlite3
    conn = sqlite3.connect(patch_db)
    conn.execute(
        "UPDATE reviews SET headline_cn = '好产品', body_cn = '非常喜欢', translate_status = 'done'"
    )
    conn.commit()
    conn.close()

    from qbu_crawler.server.report import generate_report

    since = datetime.now(timezone.utc) - timedelta(hours=1)
    result = generate_report(since, send_email=False)

    assert result["reviews_count"] == 1
    assert result["translated_count"] == 1
    assert result["untranslated_count"] == 0


def test_generate_report_with_email(patch_db, tmp_path, monkeypatch):
    """Email fails gracefully when SMTP not configured."""
    monkeypatch.setattr(config, "DB_PATH", patch_db)
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))
    monkeypatch.setattr(config, "SMTP_HOST", "")  # no SMTP
    monkeypatch.setattr(config, "EMAIL_RECIPIENTS", ["test@example.com"])

    from qbu_crawler.server.report import generate_report

    since = datetime.now(timezone.utc) - timedelta(hours=1)
    result = generate_report(since, send_email=True)

    # Should not raise
    assert result["products_count"] == 1
    assert result["email"] is not None
    assert result["email"]["success"] is False
    assert "SMTP_HOST" in result["email"]["error"]


def test_generate_report_email_keeps_legacy_subject_and_body(tmp_path, monkeypatch):
    """The daily email must preserve the original subject/body template verbatim."""
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))
    monkeypatch.setattr(
        config,
        "EMAIL_RECIPIENTS",
        ["leo.xia@forcome.com", "howard.yang@forcome.com", "chloe.tu@forcome.com"],
    )

    from qbu_crawler.server import report

    products = (
        [{"site": "basspro", "ownership": "own"} for _ in range(13)]
        + [{"site": "meatyourmaker", "ownership": "competitor"} for _ in range(14)]
        + [{"site": "waltons", "ownership": "competitor"} for _ in range(13)]
    )
    reviews = [
        {
            "rating": 2,
            "translate_status": "done",
            "headline_cn": "差评标题",
            "body_cn": "差评内容",
        }
        for _ in range(236)
    ] + [
        {
            "rating": 5,
            "translate_status": "done",
            "headline_cn": "好评标题",
            "body_cn": "好评内容",
        }
        for _ in range(2345 - 236)
    ]

    monkeypatch.setattr(report, "query_report_data", lambda since: (products, reviews))

    excel_path = tmp_path / "reports" / "scrape-report-2026-03-27.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    excel_path.write_text("stub", encoding="utf-8")
    monkeypatch.setattr(
        report,
        "generate_excel",
        lambda products, reviews, report_date=None: str(excel_path),
    )

    captured = {}

    def fake_send_email(recipients, subject, body_text, attachment_path=None):
        captured["recipients"] = recipients
        captured["subject"] = subject
        captured["body_text"] = body_text
        captured["attachment_path"] = attachment_path
        return {"success": True, "error": None, "recipients": len(recipients)}

    monkeypatch.setattr(report, "_send_email_impl", fake_send_email)

    since = datetime(2026, 3, 27, tzinfo=timezone.utc)
    result = report.generate_report(since, send_email=True)

    assert result["email"] == {"success": True, "error": None, "recipients": 3}
    assert captured["subject"] == "【网评监控】Bass Pro Shops、Meat Your Maker、waltons 产品评论报告 2026-03-27"
    assert captured["body_text"] == (
        "各位好，\n\n"
        "附件是 2026-03-27 从 Bass Pro Shops、Meat Your Maker、waltons 抓取的最新产品网评报告，请查阅。\n\n"
        "【数据概览】\n"
        "  - 涉及产品：40 个（自有 13，竞品 27）\n"
        "  - 新增评论：2345 条（已翻译 2345 条）\n"
        "\n"
        "【差评预警】共 236 条差评（≤2星），请重点关注并更新改进措施。\n"
        "\n"
        "详细数据见附件 Excel（产品 + 评论两个 Sheet）。\n"
        "如有疑问请随时沟通，谢谢！\n"
    )
    assert captured["attachment_path"] == str(excel_path)


def test_build_daily_deep_report_email_renders_incremental_summary():
    from qbu_crawler.server import report

    snapshot = {
        "logical_date": "2026-04-03",
        "data_until": "2026-04-03T23:59:59+08:00",
    }
    analytics = {
        "mode": "incremental",
        "kpis": {
            "product_count": 12,
            "own_product_count": 5,
            "competitor_product_count": 7,
            "ingested_review_rows": 186,
            "own_review_rows": 73,
            "competitor_review_rows": 113,
            "image_review_rows": 14,
            "low_rating_review_rows": 21,
            "translated_count": 180,
            "untranslated_count": 6,
        },
        "self": {
            "risk_products": [
                {
                    "product_name": "Own Grinder",
                    "product_sku": "OWN-1",
                    "negative_review_rows": 11,
                    "image_review_rows": 3,
                    "top_labels": [
                        {"label_code": "quality_stability", "count": 5},
                        {"label_code": "material_finish", "count": 3},
                    ],
                }
            ],
            "top_negative_clusters": [
                {
                    "label_code": "quality_stability",
                    "review_count": 9,
                    "image_review_count": 2,
                    "severity": "high",
                }
            ],
            "recommendations": [
                {
                    "label_code": "quality_stability",
                    "priority": "high",
                    "possible_cause_boundary": "可能与核心部件耐久性有关",
                    "improvement_direction": "优先复核高频失效部件寿命",
                    "evidence_count": 9,
                }
            ],
        },
        "competitor": {
            "top_positive_themes": [
                {
                    "label_code": "solid_build",
                    "review_count": 18,
                    "image_review_count": 2,
                },
                {
                    "label_code": "strong_performance",
                    "review_count": 12,
                    "image_review_count": 0,
                }
            ],
            "benchmark_examples": [
                {
                    "product_name": "Competitor Grinder",
                    "product_sku": "COMP-1",
                    "label_codes": ["solid_build", "strong_performance"],
                    "headline_cn": "结实耐用",
                    "body_cn": "机器做工扎实而且动力强。",
                    "headline": "Solid and strong",
                    "body": "Built well and performs great.",
                }
            ],
            "negative_opportunities": [
                {
                    "product_name": "Competitor Grinder",
                    "product_sku": "COMP-2",
                    "label_codes": ["service_fulfillment"],
                }
            ],
        },
    }

    subject, body = report.build_daily_deep_report_email(snapshot, analytics)

    assert "产品评论日报" in subject
    assert "2026-04-03" in subject
    assert "Own Grinder" in subject
    assert "今日要点：" in body
    assert "2026-04-03" in body
    assert "详见附件 PDF" in body
    assert "问题簇与改良方向" not in body


def test_generate_report_email_no_recipients(patch_db, tmp_path, monkeypatch):
    """Email is skipped gracefully when no recipients configured."""
    monkeypatch.setattr(config, "DB_PATH", patch_db)
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "reports"))
    monkeypatch.setattr(config, "SMTP_HOST", "smtp.example.com")
    monkeypatch.setattr(config, "SMTP_PORT", 587)
    monkeypatch.setattr(config, "SMTP_USER", "")
    monkeypatch.setattr(config, "SMTP_PASSWORD", "")
    monkeypatch.setattr(config, "SMTP_FROM", "")
    monkeypatch.setattr(config, "SMTP_USE_SSL", False)
    monkeypatch.setattr(config, "EMAIL_RECIPIENTS", [])

    from qbu_crawler.server.report import generate_report

    since = datetime.now(timezone.utc) - timedelta(hours=1)
    result = generate_report(since, send_email=True)

    # No recipients → send_email returns error, but generate_report doesn't raise
    assert result["email"] is not None
    assert result["email"]["success"] is False
