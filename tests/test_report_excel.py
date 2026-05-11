"""Tests for Excel workbook generation (legacy 2-sheet + parallel image utilities).

F011 §4.3 — retired: 4 sheets reorganization
─────────────────────────────────────────────
The 5-sheet analytical layout (评论明细 / 产品概览 / 今日变化 / 问题标签 /
趋势数据) was retired in F011 §4.3. The Excel attachment now produces 4
sheets (核心数据 / 行动建议 / 评论原文 / 竞品启示).

Replacement coverage for the new layout lives in
``tests/server/test_excel_sheets.py``.

This file retains the still-relevant tests:
  - Legacy 2-sheet path (no analytics arg) — kept for backward compat
  - Parallel image-download helper invariants
  - output_path copy logic
"""

import os
from datetime import datetime, timezone

import pytest

from qbu_crawler import config
from qbu_crawler.server import report


# ---------------------------------------------------------------------------
# Sample data (used by retained legacy + output_path tests)
# ---------------------------------------------------------------------------

SAMPLE_PRODUCTS = [
    {
        "url": "https://example.com/p/1",
        "name": "Test Grinder",
        "sku": "SKU-001",
        "price": 49.99,
        "stock_status": "in_stock",
        "rating": 4.2,
        "review_count": 20,
        "scraped_at": "2026-04-01 08:00:00",
        "site": "basspro",
        "ownership": "own",
    },
    {
        "url": "https://example.com/p/2",
        "name": "Rival Grinder",
        "sku": "SKU-002",
        "price": 59.99,
        "stock_status": "in_stock",
        "rating": 4.5,
        "review_count": 30,
        "scraped_at": "2026-04-01 08:00:00",
        "site": "basspro",
        "ownership": "competitor",
    },
]

SAMPLE_REVIEWS = [
    {
        "product_name": "Test Grinder",
        "product_sku": "SKU-001",
        "author": "Alice",
        "headline": "Blade broke",
        "body": "The blade broke after 2 uses.",
        "headline_cn": "刀片断了",
        "body_cn": "使用两次后刀片就断了。",
        "rating": 1.0,
        "date_published": "2026-03-28",
        "images": [],
        "ownership": "own",
        "sentiment": "negative",
        "features": ["quality_stability"],
        "insight_cn": "刀片质量问题",
    },
    {
        "product_name": "Rival Grinder",
        "product_sku": "SKU-002",
        "author": "Bob",
        "headline": "Works great",
        "body": "Solid build and easy to use.",
        "headline_cn": "很好用",
        "body_cn": "做工扎实，操作简单。",
        "rating": 5.0,
        "date_published": "2026-03-29",
        "images": ["https://img.example.com/1.jpg"],
        "ownership": "competitor",
        "sentiment": "positive",
        "features": ["solid_build", "easy_to_use"],
        "insight_cn": "",
    },
]


def _make_test_analytics():
    """Minimal valid analytics dict — used by the surviving output_path test."""
    return {
        "mode": "baseline",
        "mode_display": "首日全量基线版",
        "kpis": {},
        "self": {"risk_products": [], "product_status": []},
        "competitor": {"benchmark_examples": [], "negative_opportunities": []},
        "report_copy": {"improvement_priorities": []},
        "appendix": {},
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _patch_image_download(monkeypatch):
    """Prevent actual HTTP calls for image downloads in all tests."""
    monkeypatch.setattr(report, "_download_image_data", lambda url: None)


def test_generate_excel_without_analytics_uses_legacy(tmp_path, monkeypatch):
    """Backward compat: no analytics -> 2-sheet legacy format."""
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    path = report.generate_excel(
        products=SAMPLE_PRODUCTS,
        reviews=SAMPLE_REVIEWS,
    )
    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert "产品" in wb.sheetnames
    assert "评论" in wb.sheetnames


def test_analytical_excel_none_analytics_falls_back_to_legacy(tmp_path, monkeypatch):
    """_generate_analytical_excel with None analytics -> legacy format."""
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))
    path = report._generate_analytical_excel(
        SAMPLE_PRODUCTS, SAMPLE_REVIEWS, analytics=None,
    )
    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert "产品" in wb.sheetnames
    assert "评论" in wb.sheetnames


def test_generate_excel_output_path_with_analytics(tmp_path, monkeypatch):
    """Output path copy logic works with the analytical format."""
    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path / "default"))
    custom_out = str(tmp_path / "custom" / "report.xlsx")
    analytics = _make_test_analytics()

    path = report.generate_excel(
        SAMPLE_PRODUCTS, SAMPLE_REVIEWS,
        analytics=analytics,
        output_path=custom_out,
    )
    assert path == custom_out
    assert os.path.isfile(custom_out)


def test_download_images_parallel_respects_timeout(monkeypatch):
    """Parallel image downloads should complete within global timeout."""
    import time
    from qbu_crawler.server.report import _download_images_parallel

    call_count = 0

    def slow_download(url, timeout=10):
        nonlocal call_count
        call_count += 1
        time.sleep(0.3)
        return None  # Simulate failed download

    monkeypatch.setattr("qbu_crawler.server.report._download_image_data", slow_download)
    urls = [f"https://img.example.com/{i}.jpg" for i in range(10)]
    start = time.time()
    results = _download_images_parallel(urls, global_timeout=3)
    elapsed = time.time() - start

    assert len(results) == 10
    # With 5 workers and 0.3s per download, 10 items should take ~0.6s (2 rounds)
    # Certainly less than serial 3.0s
    assert elapsed < 2.0, f"Took too long: {elapsed:.1f}s"


def test_download_images_parallel_empty_list():
    """Empty URL list should return empty results."""
    from qbu_crawler.server.report import _download_images_parallel

    results = _download_images_parallel([], global_timeout=5)
    assert results == {}


def test_generate_excel_calls_parallel_prefetch(monkeypatch, tmp_path):
    """Legacy generate_excel should call _download_images_parallel for pre-fetching."""
    from qbu_crawler.server import report

    monkeypatch.setattr(config, "REPORT_DIR", str(tmp_path))

    prefetch_calls = []

    def mock_parallel(urls, **kwargs):
        prefetch_calls.append(urls)
        return {url: None for url in urls}

    monkeypatch.setattr(report, "_download_images_parallel", mock_parallel)
    monkeypatch.setattr(report, "_download_image_data", lambda url: None)

    reviews_with_images = [
        {
            "product_name": "P1",
            "product_sku": "SKU1",
            "author": "A1",
            "headline": "Great",
            "body": "Body",
            "headline_cn": "",
            "body_cn": "",
            "rating": 5,
            "date_published": "2026-01-01",
            "images": ["https://img.example.com/a.jpg", "https://img.example.com/b.jpg"],
        },
    ]
    report._legacy_generate_excel([], reviews_with_images)

    assert len(prefetch_calls) == 1
    assert set(prefetch_calls[0]) == {"https://img.example.com/a.jpg", "https://img.example.com/b.jpg"}
